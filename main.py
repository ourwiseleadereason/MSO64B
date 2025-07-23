# date: 2024-07-08
# author: Grace Qian
# Final version for I2C test report
# store csv file in test data/I2C/csv
# store capture files in test data/I2C/waveform
# store test plan in "test plan" folder
# change test plan name in line 134
# generate test report named as "I2C test report.xlsx" in the root folder


import os
import openpyxl
import pandas as pd
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import colors
from openpyxl.drawing.image import Image
from openpyxl.drawing.xdr import XDRPoint2D, XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU, cm_to_EMU


def csv_handle_new(path, target_row, target_col):

    largest_column_count = 0

    with open(path, 'r') as f:
        # reader = csv.reader(f)
        lines = f.readlines()
        # print(lines)
        for line in lines:
            column_count = len(line.split(',')) + 1
            largest_column_count = column_count if largest_column_count < column_count else largest_column_count
    f.close()
    # print(largest_column_count)
    column_names = [i for i in range(0, largest_column_count)]

    # df = pd.read_csv(path, header=None, delimiter=',', names=column_names)
    df = pd.read_csv(path, header=None, delimiter=',')

    # row_count = df.shape[0]
    # col_count = df.shape[1]
    # print(row_count, col_count)

    # for row in range(0, int(row_count)):
    #     for col in range(0, int(col_count)):
    #
    #         if df[col][row] == target_row:
    #             row_target = row
    #             # print(row_target)
    #         if df[col][row] == target_col:
    #             col_target = col
    # # print(row_target, col_target)
    mea_data = df[target_col][target_row]
    return mea_data


def get_freq(path):

    freq_t = round(float(csv_handle_new(path, 0, 1) / 1000), 3)
    high_T_t = round(float(csv_handle_new(path, 1, 1) * 1e6), 3)
    low_T_t = round(float(csv_handle_new(path, 2, 1) * 1e6), 3)
    return freq_t, high_T_t, low_T_t


def get_SCL_SDA_voltage(path):

    SCL_Max_v = round(float(csv_handle_new(path, 0, 1)), 3)
    SCL_Min_v = round(float(csv_handle_new(path, 1, 1)), 3)
    SCL_H_v = round(float(csv_handle_new(path, 2, 1)), 3)
    SCL_L_v = round(float(csv_handle_new(path, 3, 1)), 3)
    SDA_Max_v = round(float(csv_handle_new(path, 4, 1)), 3)
    SDA_Min_v = round(float(csv_handle_new(path, 5, 1)), 3)
    SDA_H_v = round(float(csv_handle_new(path, 6, 1)), 3)
    SDA_L_v = round(float(csv_handle_new(path, 7, 1)), 3)
    return SCL_Max_v, SCL_Min_v, SCL_H_v, SCL_L_v, SDA_Max_v, SDA_Min_v, SDA_H_v, SDA_L_v


def get_SCL_SDA_RF(path):

    SCL_R_t = round(float(csv_handle_new(path, 0, 1) * 1e9), 3)
    SCL_F_t = round(float(csv_handle_new(path, 1, 1) * 1e9), 3)
    SDA_R_t = round(float(csv_handle_new(path, 2, 1) * 1e9), 3)
    SDA_F_t = round(float(csv_handle_new(path, 3, 1) * 1e9), 3)
    return SCL_R_t, SCL_F_t, SDA_R_t, SDA_F_t


def get_tsu_thd(path):

    tsu_thd_t = round(float(csv_handle_new(path, 0, 1)*1e9), 3)
    return tsu_thd_t


# get start and stop time
def get_sta_sto(path):

    sta_sto_t = round(float(csv_handle_new(path, 0, 1)*1e6), 3)
    return sta_sto_t


def get_SDA_voltage(path):

    SDA_Max_v = round(float(csv_handle_new(path, 0, 1)), 3)
    SDA_Min_v = round(float(csv_handle_new(path, 1, 1)), 3)
    SDA_H_v = round(float(csv_handle_new(path, 2, 1)), 3)
    SDA_L_v = round(float(csv_handle_new(path, 3, 1)), 3)
    return SDA_Max_v, SDA_Min_v, SDA_H_v, SDA_L_v


def get_SDA_RF(path):

    SDA_R_t = round(float(csv_handle_new(path, 0, 1) * 1e9), 3)
    SDA_F_t = round(float(csv_handle_new(path, 1, 1) * 1e9), 3)
    return SDA_R_t, SDA_F_t


def inserImg(wsheet, Imgpath, position):

    imgToInsert = Image(Imgpath)
    oriImgH, oriImgW = imgToInsert.height, imgToInsert.width

    resize_factor = 0.25
    w_h_raio = oriImgW/oriImgH
    resize_H = resize_factor * oriImgH
    resize_W = resize_factor * oriImgW
    imgsize_excel = XDRPositiveSize2D(pixels_to_EMU(resize_W), pixels_to_EMU(resize_H))
    imgToInsert.height = resize_H
    imgToInsert.width = resize_W
    imgToInsert.anchor = position
    wsheet.add_image(imgToInsert)


fill_yellow = PatternFill('solid', fgColor='f8c600')  # Yellow color
fill_red = PatternFill('solid', fgColor='ff0000')  # Red color
fill_green = PatternFill('solid', fgColor='00ff00')  # Green color
font_blue = Font(color=colors.BLUE)  # blue color
font_red = Font(color='FF0000')  # red color

wb = openpyxl.load_workbook('1 [TPDR-05-1-001]R0B I2C test report_v0.2.xlsx')
ws = wb['I2C with ELB']
rows = ws.max_row
cols = ws.max_column

root = os.getcwd()
csv_path = os.path.join(root, 'test data', 'I2C', 'csv')
id_list = os.listdir(csv_path)
cap_path = os.path.join(root, 'test data', 'I2C', 'waveform')
print(cap_path)
print(id_list)
print(len(id_list))

for idn in id_list:

    idn_path = os.path.join(csv_path, idn)
    item_list = os.listdir(idn_path)
    num_csv = len(item_list)
    print(idn, num_csv, item_list)

    if int(num_csv) == 7:
        f_Path = os.path.join(idn_path, 'freq.csv')
        freq, high_T, low_T = get_freq(f_Path)
        v_Path = os.path.join(idn_path, 'v.csv')
        SCL_Max, SCL_Min, SCL_H, SCL_L, SDA_Max, SDA_Min, SDA_H, SDA_L = get_SCL_SDA_voltage(v_Path)
        rf_Path = os.path.join(idn_path, 'rf.csv')
        SCL_R, SCL_F, SDA_R, SDA_F = get_SCL_SDA_RF(rf_Path)
        tsu_Path = os.path.join(idn_path, 'tsu.csv')
        tsu_mea = get_tsu_thd(tsu_Path)
        thd_Path = os.path.join(idn_path, 'thd.csv')
        thd_mea = get_tsu_thd(thd_Path)
        sta_Path = os.path.join(idn_path, 'sta.csv')
        sta_mea = get_sta_sto(sta_Path)
        sto_Path = os.path.join(idn_path, 'sto.csv')
        sto_mea = get_sta_sto(sto_Path)

    if int(num_csv) == 1:
        v_Path = os.path.join(idn_path, 'v.csv')
        # SCL_Max, SCL_Min, SCL_H, SCL_L, SDA_Max, SDA_Min, SDA_H, SDA_L = get_SCL_SDA_voltage(v_Path)
        SCL_Max, SCL_Min, SCL_H, SCL_L, SDA_Max, SDA_Min, SDA_H, SDA_L = get_SCL_SDA_voltage(v_Path)

    if int(num_csv) == 6:
        v_Path = os.path.join(idn_path, 'v.csv')
        SDA_Max, SDA_Min, SDA_H, SDA_L = get_SDA_voltage(v_Path)
        rf_Path = os.path.join(idn_path, 'rf.csv')
        SDA_R, SDA_F = get_SDA_RF(rf_Path)
        tsu_Path = os.path.join(idn_path, 'tsu.csv')
        tsu_mea = get_tsu_thd(tsu_Path)
        thd_Path = os.path.join(idn_path, 'thd.csv')
        thd_mea = get_tsu_thd(thd_Path)
        sta_Path = os.path.join(idn_path, 'sta.csv')
        sta_mea = get_sta_sto(sta_Path)
        sto_Path = os.path.join(idn_path, 'sto.csv')
        sto_mea = get_sta_sto(sto_Path)

    tar_id = 'Sub-case ID:' + idn
    tar_col = 5
    mar_col = 7
    res_col = 8

    # print(tar_id)
    for row in range(1, rows+1):
        column = 2
        value = ws.cell(row, column).value
        # print(value)
        if tar_id == str(value):
            id_row = row
            print(id_row)

    # write data to test plan
    if int(num_csv) == 7:

        # input the timing value to the test plan
        ws.cell(id_row + 6, tar_col).value = freq
        ws.cell(id_row + 7, tar_col).value = high_T
        ws.cell(id_row + 8, tar_col).value = low_T

        # input the voltage value to the test plan
        ws.cell(id_row + 9, tar_col).value = SCL_Max
        ws.cell(id_row + 10, tar_col).value = SCL_Min
        ws.cell(id_row + 11, tar_col).value = SCL_H
        ws.cell(id_row + 12, tar_col).value = SCL_L

        ws.cell(id_row + 13, tar_col).value = SDA_Max
        ws.cell(id_row + 14, tar_col).value = SDA_Min
        ws.cell(id_row + 15, tar_col).value = SDA_H
        ws.cell(id_row + 16, tar_col).value = SDA_L

        # input rise time and fall time
        ws.cell(id_row + 17, tar_col).value = SCL_R
        ws.cell(id_row + 18, tar_col).value = SCL_F
        ws.cell(id_row + 19, tar_col).value = SDA_R
        ws.cell(id_row + 20, tar_col).value = SDA_F

        # input SDA setup time and hold time
        ws.cell(id_row + 21, tar_col).value = tsu_mea
        ws.cell(id_row + 22, tar_col).value = thd_mea

        # input start and stop time
        ws.cell(id_row + 23, tar_col).value = sta_mea
        ws.cell(id_row + 24, tar_col).value = sto_mea

        # SCL Monotonic
        ws.cell(id_row + 25, tar_col).value = 'Yes'

        # wb.save('I2C test report.xlsx')

    # write data to test plan
    if int(num_csv) == 1:

        # input the voltage value to the test plan
        ws.cell(id_row + 6, tar_col).value = SCL_Max
        ws.cell(id_row + 7, tar_col).value = SCL_Min
        ws.cell(id_row + 8, tar_col).value = SCL_H
        ws.cell(id_row + 9, tar_col).value = SCL_L

        ws.cell(id_row + 10, tar_col).value = SDA_Max
        ws.cell(id_row + 11, tar_col).value = SDA_Min
        ws.cell(id_row + 12, tar_col).value = SDA_H
        ws.cell(id_row + 13, tar_col).value = SDA_L

        # wb.save('I2C test report.xlsx')

    # write data to test plan
    if int(num_csv) == 6:

        ws.cell(id_row + 6, tar_col).value = SDA_Max
        ws.cell(id_row + 7, tar_col).value = SDA_Min
        ws.cell(id_row + 8, tar_col).value = SDA_H
        ws.cell(id_row + 9, tar_col).value = SDA_L

        ws.cell(id_row + 10, tar_col).value = SDA_R
        ws.cell(id_row + 11, tar_col).value = SDA_F

        ws.cell(id_row + 12, tar_col).value = tsu_mea
        ws.cell(id_row + 13, tar_col).value = thd_mea

        ws.cell(id_row + 14, tar_col).value = sta_mea
        ws.cell(id_row + 15, tar_col).value = sto_mea

        # wb.save('I2C test report.xlsx')

    # Margin calculation
    if int(num_csv) == 7:

        # frequency/high time/low time margin
        if ws.cell(id_row + 6, tar_col - 2).value < ws.cell(id_row + 6, tar_col).value < ws.cell(id_row + 6, tar_col - 1).value:
            ws.cell(id_row + 6, tar_col).font = font_blue
            ws.cell(id_row + 6, mar_col).value = '/'
            ws.cell(id_row + 6, mar_col).font = font_blue
            ws.cell(id_row + 6, res_col).value = 'PASS'
            ws.cell(id_row + 6, res_col).fill = fill_green
        else:
            ws.cell(id_row + 6, tar_col).fill = fill_red
            # ws.cell(id_row + 6, mar_col).value = '/'
            # ws.cell(id_row + 6, mar_col).font = font_blue
            ws.cell(id_row + 6, res_col).value = 'FAIL'
            ws.cell(id_row + 6, res_col).fill = fill_red

        # high time margin
        if ws.cell(id_row + 7, tar_col).value > ws.cell(id_row + 7, tar_col - 2).value:
            ws.cell(id_row + 7, tar_col).font = font_blue
            ws.cell(id_row + 7, mar_col).value = \
                (ws.cell(id_row + 7, tar_col).value - ws.cell(id_row + 7, tar_col - 2).value) / \
                ws.cell(id_row + 7, tar_col - 2).value
            ws.cell(id_row + 7, mar_col).number_format = '0%'
            # set the cell color and result
            if ws.cell(id_row + 7, mar_col).value >= 0.05:
                ws.cell(id_row + 7, tar_col).font = font_blue
                ws.cell(id_row + 7, mar_col).font = font_blue
                ws.cell(id_row + 7, res_col).value = 'PASS'
                ws.cell(id_row + 7, res_col).fill = fill_green
            if 0 < ws.cell(id_row + 7, mar_col).value < 0.05:
                ws.cell(id_row + 7, tar_col).fill = fill_yellow
                ws.cell(id_row + 7, mar_col).fill = fill_yellow
                ws.cell(id_row + 7, res_col).value = 'Marginal Pass'
                ws.cell(id_row + 7, res_col).fill = fill_yellow
        else:
            ws.cell(id_row + 7, tar_col).fill = fill_red
            ws.cell(id_row + 7, res_col).value = 'FAIL'
            ws.cell(id_row + 7, res_col).fill = fill_red

        # low time margin
        if ws.cell(id_row + 8, tar_col).value > ws.cell(id_row + 8, tar_col - 2).value:
            ws.cell(id_row + 8, tar_col).font = font_blue
            ws.cell(id_row + 8, mar_col).value = \
                (ws.cell(id_row + 8, tar_col).value - ws.cell(id_row + 8, tar_col - 2).value) / \
                ws.cell(id_row + 8, tar_col - 2).value
            ws.cell(id_row + 8, mar_col).number_format = '0%'
            # set the cell color and result
            if ws.cell(id_row + 8, mar_col).value >= 0.05:
                ws.cell(id_row + 8, tar_col).font = font_blue
                ws.cell(id_row + 8, mar_col).font = font_blue
                ws.cell(id_row + 8, res_col).value = 'PASS'
                ws.cell(id_row + 8, res_col).fill = fill_green
            if 0 < ws.cell(id_row + 8, mar_col).value < 0.05:
                ws.cell(id_row + 8, tar_col).fill = fill_yellow
                ws.cell(id_row + 8, mar_col).fill = fill_yellow
                ws.cell(id_row + 8, res_col).value = 'Marginal Pass'
                ws.cell(id_row + 8, res_col).fill = fill_yellow
        else:
            ws.cell(id_row + 8, tar_col).fill = fill_red
            ws.cell(id_row + 8, res_col).value = 'FAIL'
            ws.cell(id_row + 8, res_col).fill = fill_red

        # SCL_MAX margin
        if type(ws.cell(id_row + 9, tar_col - 1).value) is str:
            if ws.cell(id_row + 9, tar_col - 1).value == 'NA' or '/':

                ws.cell(id_row + 9, tar_col).font = font_blue
                ws.cell(id_row + 9, mar_col).value = '/'
                ws.cell(id_row + 9, res_col).value = 'For Reference'

        else:
            if ws.cell(id_row + 9, tar_col).value < ws.cell(id_row + 9, tar_col - 1).value:

                ws.cell(id_row + 9, mar_col).value = \
                    (ws.cell(id_row + 9, tar_col-1).value - ws.cell(id_row + 9, tar_col).value)/\
                    ws.cell(id_row + 9, tar_col-1).value
                ws.cell(id_row + 9, mar_col).number_format = '0%'

                if ws.cell(id_row + 9, mar_col).value >= 0.05:
                    ws.cell(id_row + 9, tar_col).font = font_blue
                    ws.cell(id_row + 9, mar_col).font = font_blue
                    ws.cell(id_row + 9, res_col).value = 'PASS'
                    ws.cell(id_row + 9, res_col).fill = fill_green

                if 0 < ws.cell(id_row + 9, mar_col).value < 0.05:
                    ws.cell(id_row + 9, tar_col).fill = fill_yellow
                    ws.cell(id_row + 9, mar_col).fill = fill_yellow
                    ws.cell(id_row + 9, res_col).value = 'Marginal Pass'
                    ws.cell(id_row + 9, res_col).fill = fill_yellow
            else:
                ws.cell(id_row + 9, tar_col).fill = fill_red
                # ws.cell(id_row + 9, mar_col).fill = fill_yellow
                ws.cell(id_row + 9, res_col).value = 'FAIL'
                ws.cell(id_row + 9, res_col).fill = fill_red

        # SCL_Min margin
        if type(ws.cell(id_row + 10, tar_col - 2).value) is str:
            if ws.cell(id_row + 10, tar_col - 2).value == 'NA' or '/':
                ws.cell(id_row + 10, tar_col).font = font_blue
                ws.cell(id_row + 10, mar_col).value = '/'
                ws.cell(id_row + 10, res_col).value = 'For Reference'

        else:
            if ws.cell(id_row + 10, tar_col).value > ws.cell(id_row + 10, tar_col - 2).value:

                ws.cell(id_row + 10, mar_col).value = \
                    (ws.cell(id_row + 10, tar_col - 2).value - ws.cell(id_row + 10, tar_col).value) / \
                    ws.cell(id_row + 10, tar_col - 2).value
                ws.cell(id_row + 10, mar_col).number_format = '0%'

                if ws.cell(id_row + 10, mar_col).value >= 0.05:
                    ws.cell(id_row + 10, tar_col).font = font_blue
                    ws.cell(id_row + 10, mar_col).font = font_blue
                    ws.cell(id_row + 10, res_col).value = 'PASS'
                    ws.cell(id_row + 10, res_col).fill = fill_green

                if 0 < ws.cell(id_row + 10, mar_col).value < 0.05:
                    ws.cell(id_row + 10, tar_col).fill = fill_yellow
                    ws.cell(id_row + 10, mar_col).fill = fill_yellow
                    ws.cell(id_row + 10, res_col).value = 'Marginal Pass'
                    ws.cell(id_row + 10, res_col).fill = fill_yellow
            else:
                ws.cell(id_row + 10, tar_col).fill = fill_red
                # ws.cell(id_row + 9, mar_col).fill = fill_yellow
                ws.cell(id_row + 10, res_col).value = 'FAIL'
                ws.cell(id_row + 10, res_col).fill = fill_red

        # SCL_H margin
        if type(ws.cell(id_row+11, tar_col-2).value) is str:
            if ws.cell(id_row+11, tar_col-2).value == 'NA' or '/':
                ws.cell(id_row + 11, tar_col).font = font_blue
                ws.cell(id_row + 11, mar_col).value = '/'
                ws.cell(id_row + 11, res_col).value = 'For Reference'

        else:
            if ws.cell(id_row+11, tar_col).value > ws.cell(id_row+11, tar_col-2).value:
                ws.cell(id_row+11, mar_col).value = \
                    (ws.cell(id_row+11, tar_col).value - ws.cell(id_row+11, tar_col-2).value)/\
                    ws.cell(id_row+11, tar_col-2).value
                ws.cell(id_row+11, mar_col).number_format = '0%'

                if ws.cell(id_row + 11, mar_col).value >= 0.05:
                    ws.cell(id_row + 11, tar_col).font = font_blue
                    ws.cell(id_row + 11, mar_col).font = font_blue
                    ws.cell(id_row + 11, res_col).value = 'PASS'
                    ws.cell(id_row + 11, res_col).fill = fill_green

                if 0 < ws.cell(id_row + 11, mar_col).value < 0.05:
                    ws.cell(id_row + 11, tar_col).fill = fill_yellow
                    ws.cell(id_row + 11, mar_col).fill = fill_yellow
                    ws.cell(id_row + 11, res_col).value = 'Marginal Pass'
                    ws.cell(id_row + 11, res_col).fill = fill_yellow
            else:
                ws.cell(id_row + 11, tar_col).fill = fill_red
                ws.cell(id_row + 11, res_col).value = 'FAIL'
                ws.cell(id_row + 11, res_col).fill = fill_red

        # SCL_L margin
        if type(ws.cell(id_row+12, tar_col-1).value) is str:
            if ws.cell(id_row+12, tar_col-1).value == 'NA' or '/':
                ws.cell(id_row + 12, tar_col).font = font_blue
                ws.cell(id_row + 12, mar_col).value = '/'
                ws.cell(id_row + 12, res_col).value = 'For Reference'

        else:
            if ws.cell(id_row + 12, tar_col).value < ws.cell(id_row + 12, tar_col - 1).value:
                ws.cell(id_row + 12, mar_col).value = \
                    (ws.cell(id_row + 12, tar_col-1).value - ws.cell(id_row + 12, tar_col).value) / \
                    ws.cell(id_row + 12, tar_col-1).value
                ws.cell(id_row + 12, mar_col).number_format = '0%'

                if ws.cell(id_row + 12, mar_col).value >= 0.05:
                    ws.cell(id_row + 12, tar_col).font = font_blue
                    ws.cell(id_row + 12, mar_col).font = font_blue
                    ws.cell(id_row + 12, res_col).value = 'PASS'
                    ws.cell(id_row + 12, res_col).fill = fill_green

                if 0 < ws.cell(id_row + 12, mar_col).value < 0.05:
                    ws.cell(id_row + 12, tar_col).fill = fill_yellow
                    ws.cell(id_row + 12, mar_col).fill = fill_yellow
                    ws.cell(id_row + 12, res_col).value = 'Marginal Pass'
                    ws.cell(id_row + 12, res_col).fill = fill_yellow
            else:
                ws.cell(id_row + 12, tar_col).fill = fill_red
                ws.cell(id_row + 12, res_col).value = 'FAIL'
                ws.cell(id_row + 12, res_col).fill = fill_red

        # SDA_MAX margin
        if type(ws.cell(id_row + 13, tar_col - 1).value) is str:
            if ws.cell(id_row + 13, tar_col - 1).value == 'NA' or '/':

                ws.cell(id_row + 13, tar_col).font = font_blue
                ws.cell(id_row + 13, mar_col).value = '/'
                ws.cell(id_row + 13, res_col).value = 'For Reference'

        else:
            if ws.cell(id_row + 13, tar_col).value < ws.cell(id_row + 13, tar_col - 1).value:

                ws.cell(id_row + 13, mar_col).value = \
                    (ws.cell(id_row + 13, tar_col - 1).value - ws.cell(id_row + 13, tar_col).value) / \
                    ws.cell(id_row + 13, tar_col - 1).value
                ws.cell(id_row + 13, mar_col).number_format = '0%'

                if ws.cell(id_row + 13, mar_col).value >= 0.05:
                    ws.cell(id_row + 13, tar_col).font = font_blue
                    ws.cell(id_row + 13, mar_col).font = font_blue
                    ws.cell(id_row + 13, res_col).value = 'PASS'
                    ws.cell(id_row + 13, res_col).fill = fill_green

                if 0 < ws.cell(id_row + 13, mar_col).value < 0.05:
                    ws.cell(id_row + 13, tar_col).fill = fill_yellow
                    ws.cell(id_row + 13, mar_col).fill = fill_yellow
                    ws.cell(id_row + 13, res_col).value = 'Marginal Pass'
                    ws.cell(id_row + 13, res_col).fill = fill_yellow
            else:
                ws.cell(id_row + 13, tar_col).fill = fill_red
                # ws.cell(id_row + 9, mar_col).fill = fill_yellow
                ws.cell(id_row + 13, res_col).value = 'FAIL'
                ws.cell(id_row + 13, res_col).fill = fill_red

        # SDA_Min margin
        if type(ws.cell(id_row + 14, tar_col - 2).value) is str:
            if ws.cell(id_row + 14, tar_col - 2).value == 'NA' or '/':
                ws.cell(id_row + 14, tar_col).font = font_blue
                ws.cell(id_row + 14, mar_col).value = '/'
                ws.cell(id_row + 14, res_col).value = 'For Reference'

        else:
            if ws.cell(id_row + 14, tar_col).value > ws.cell(id_row + 14, tar_col - 2).value:

                ws.cell(id_row + 14, mar_col).value = \
                    (ws.cell(id_row + 14, tar_col - 2).value - ws.cell(id_row + 14, tar_col).value) / \
                    ws.cell(id_row + 14, tar_col - 2).value
                ws.cell(id_row + 14, mar_col).number_format = '0%'

                if ws.cell(id_row + 14, mar_col).value >= 0.05:
                    ws.cell(id_row + 14, tar_col).font = font_blue
                    ws.cell(id_row + 14, mar_col).font = font_blue
                    ws.cell(id_row + 14, res_col).value = 'PASS'
                    ws.cell(id_row + 14, res_col).fill = fill_green

                if 0 < ws.cell(id_row + 14, mar_col).value < 0.05:
                    ws.cell(id_row + 14, tar_col).fill = fill_yellow
                    ws.cell(id_row + 14, mar_col).fill = fill_yellow
                    ws.cell(id_row + 14, res_col).value = 'Marginal Pass'
                    ws.cell(id_row + 14, res_col).fill = fill_yellow
            else:
                ws.cell(id_row + 14, tar_col).fill = fill_red
                # ws.cell(id_row + 9, mar_col).fill = fill_yellow
                ws.cell(id_row + 14, res_col).value = 'FAIL'
                ws.cell(id_row + 14, res_col).fill = fill_red

        # SDA_H margin
        if type(ws.cell(id_row + 15, tar_col - 2).value) is str:
            if ws.cell(id_row + 15, tar_col - 2).value == 'NA' or '/':
                ws.cell(id_row + 15, tar_col).font = font_blue
                ws.cell(id_row + 15, mar_col).value = '/'
                ws.cell(id_row + 15, res_col).value = 'For Reference'
        else:
            if ws.cell(id_row + 15, tar_col).value > ws.cell(id_row + 15, tar_col - 2).value:
                ws.cell(id_row + 15, mar_col).value = \
                    (ws.cell(id_row + 15, tar_col).value - ws.cell(id_row + 15, tar_col - 2).value) / \
                    ws.cell(id_row + 15, tar_col - 2).value
                ws.cell(id_row + 15, mar_col).number_format = '0%'

                if ws.cell(id_row + 15, mar_col).value >= 0.05:
                    ws.cell(id_row + 15, tar_col).font = font_blue
                    ws.cell(id_row + 15, mar_col).font = font_blue
                    ws.cell(id_row + 15, res_col).value = 'PASS'
                    ws.cell(id_row + 15, res_col).fill = fill_green

                if 0 < ws.cell(id_row + 15, mar_col).value < 0.05:
                    ws.cell(id_row + 15, tar_col).fill = fill_yellow
                    ws.cell(id_row + 15, mar_col).fill = fill_yellow
                    ws.cell(id_row + 15, res_col).value = 'Marginal Pass'
                    ws.cell(id_row + 15, res_col).fill = fill_yellow
            else:
                ws.cell(id_row + 15, tar_col).fill = fill_red
                ws.cell(id_row + 15, res_col).value = 'FAIL'
                ws.cell(id_row + 15, res_col).fill = fill_red

        # SDA_L margin
        if type(ws.cell(id_row + 16, tar_col - 1).value) is str:
            if ws.cell(id_row + 16, tar_col - 1).value == 'NA' or '/':
                ws.cell(id_row + 16, tar_col).font = font_blue
                ws.cell(id_row + 16, mar_col).value = '/'
                ws.cell(id_row + 16, res_col).value = 'For Reference'

        else:
            if ws.cell(id_row + 16, tar_col).value < ws.cell(id_row + 16, tar_col - 1).value:
                ws.cell(id_row + 16, mar_col).value = \
                    (ws.cell(id_row + 16, tar_col - 1).value - ws.cell(id_row + 16, tar_col).value) / \
                    ws.cell(id_row + 16, tar_col - 1).value
                ws.cell(id_row + 16, mar_col).number_format = '0%'

                if ws.cell(id_row + 16, mar_col).value >= 0.05:
                    ws.cell(id_row + 16, tar_col).font = font_blue
                    ws.cell(id_row + 16, mar_col).font = font_blue
                    ws.cell(id_row + 16, res_col).value = 'PASS'
                    ws.cell(id_row + 16, res_col).fill = fill_green

                if 0 < ws.cell(id_row + 16, mar_col).value < 0.05:
                    ws.cell(id_row + 16, tar_col).fill = fill_yellow
                    ws.cell(id_row + 16, mar_col).fill = fill_yellow
                    ws.cell(id_row + 16, res_col).value = 'Marginal Pass'
                    ws.cell(id_row + 16, res_col).fill = fill_yellow
            else:
                ws.cell(id_row + 16, tar_col).fill = fill_red
                ws.cell(id_row + 16, res_col).value = 'FAIL'
                ws.cell(id_row + 16, res_col).fill = fill_red

        # SCL_Rise margin
        if ws.cell(id_row+17, tar_col-2).value < ws.cell(id_row+17, tar_col).value < ws.cell(id_row+17, tar_col-1).value:
            min_SCL_R = ws.cell(id_row+17, tar_col).value - ws.cell(id_row+17, tar_col-2).value
            max_SCL_R = ws.cell(id_row+17, tar_col-1).value - ws.cell(id_row+17, tar_col).value
            if max_SCL_R < min_SCL_R:
                ws.cell(id_row+17, mar_col).value = max_SCL_R / 280
            else:
                ws.cell(id_row + 17, mar_col).value = min_SCL_R / 280

            ws.cell(id_row + 17, mar_col).number_format = '0%'

            if ws.cell(id_row + 17, mar_col).value >= 0.05:
                ws.cell(id_row + 17, tar_col).font = font_blue
                ws.cell(id_row + 17, mar_col).font = font_blue
                ws.cell(id_row + 17, res_col).value = 'PASS'
                ws.cell(id_row + 17, res_col).fill = fill_green

            if 0 < ws.cell(id_row + 17, mar_col).value < 0.05:
                ws.cell(id_row + 17, tar_col).fill = fill_yellow
                ws.cell(id_row + 17, mar_col).fill = fill_yellow
                ws.cell(id_row + 17, res_col).value = 'Marginal Pass'
                ws.cell(id_row + 17, res_col).fill = fill_yellow
        else:
            ws.cell(id_row + 17, tar_col).fill = fill_red
            ws.cell(id_row + 17, res_col).value = 'FAIL'
            ws.cell(id_row + 17, res_col).fill = fill_red

        # SCL_Fall margin
        if ws.cell(id_row+18, tar_col-2).value < ws.cell(id_row+18, tar_col).value < ws.cell(id_row+18, tar_col-1).value:
            ws.cell(id_row+18, mar_col).value = 1 - ws.cell(id_row+18, tar_col).value / ws.cell(id_row+18, tar_col-1).value

            ws.cell(id_row + 18, mar_col).number_format = '0%'

            if ws.cell(id_row + 18, mar_col).value >= 0.05:
                ws.cell(id_row + 18, tar_col).font = font_blue
                ws.cell(id_row + 18, mar_col).font = font_blue
                ws.cell(id_row + 18, res_col).value = 'PASS'
                ws.cell(id_row + 18, res_col).fill = fill_green

            if 0 < ws.cell(id_row + 18, mar_col).value < 0.05:
                ws.cell(id_row + 18, tar_col).fill = fill_yellow
                ws.cell(id_row + 18, mar_col).fill = fill_yellow
                ws.cell(id_row + 18, res_col).value = 'Marginal Pass'
                ws.cell(id_row + 18, res_col).fill = fill_yellow
        else:
            ws.cell(id_row + 18, tar_col).fill = fill_red
            ws.cell(id_row + 18, res_col).value = 'FAIL'
            ws.cell(id_row + 18, res_col).fill = fill_red

        # SDA_Rise margin
        if ws.cell(id_row+19, tar_col-2).value < ws.cell(id_row+19, tar_col).value < ws.cell(id_row+17, tar_col-1).value:
            min_SCL_R = ws.cell(id_row+19, tar_col).value - ws.cell(id_row+19, tar_col-2).value
            max_SCL_R = ws.cell(id_row+19, tar_col-1).value - ws.cell(id_row+19, tar_col).value
            if max_SCL_R < min_SCL_R:
                ws.cell(id_row + 19, mar_col).value = max_SCL_R / 280
            else:
                ws.cell(id_row + 19, mar_col).value = min_SCL_R / 280

            ws.cell(id_row + 19, mar_col).number_format = '0%'

            if ws.cell(id_row + 19, mar_col).value >= 0.05:
                ws.cell(id_row + 19, tar_col).font = font_blue
                ws.cell(id_row + 19, mar_col).font = font_blue
                ws.cell(id_row + 19, res_col).value = 'PASS'
                ws.cell(id_row + 19, res_col).fill = fill_green

            if 0 < ws.cell(id_row + 19, mar_col).value < 0.05:
                ws.cell(id_row + 19, tar_col).fill = fill_yellow
                ws.cell(id_row + 19, mar_col).fill = fill_yellow
                ws.cell(id_row + 19, res_col).value = 'Marginal Pass'
                ws.cell(id_row + 19, res_col).fill = fill_yellow
        else:
            ws.cell(id_row + 19, tar_col).fill = fill_red
            ws.cell(id_row + 19, res_col).value = 'FAIL'
            ws.cell(id_row + 19, res_col).fill = fill_red

        # SDA_Fall margin
        if ws.cell(id_row+20, tar_col-2).value < ws.cell(id_row+20, tar_col).value < ws.cell(id_row+20, tar_col-1).value:
            ws.cell(id_row+20, mar_col).value = 1 - ws.cell(id_row+20, tar_col).value / ws.cell(id_row+20, tar_col-1).value

            ws.cell(id_row + 20, mar_col).number_format = '0%'

            if ws.cell(id_row + 20, mar_col).value >= 0.05:
                ws.cell(id_row + 20, tar_col).font = font_blue
                ws.cell(id_row + 20, mar_col).font = font_blue
                ws.cell(id_row + 20, res_col).value = 'PASS'
                ws.cell(id_row + 20, res_col).fill = fill_green

            if 0 < ws.cell(id_row + 20, mar_col).value < 0.05:
                ws.cell(id_row + 20, tar_col).fill = fill_yellow
                ws.cell(id_row + 20, mar_col).fill = fill_yellow
                ws.cell(id_row + 20, res_col).value = 'Marginal Pass'
                ws.cell(id_row + 20, res_col).fill = fill_yellow
        else:
            ws.cell(id_row + 20, tar_col).fill = fill_red
            ws.cell(id_row + 20, res_col).value = 'FAIL'
            ws.cell(id_row + 20, res_col).fill = fill_red

        # Tsu margin
        if ws.cell(id_row + 21, tar_col).value > ws.cell(id_row + 21, tar_col-2).value:
            ws.cell(id_row+21, mar_col).value  = \
                (ws.cell(id_row + 21, tar_col).value - ws.cell(id_row + 21, tar_col-2).value) / \
                ws.cell(id_row + 21, tar_col-2).value

            ws.cell(id_row + 21, mar_col).number_format = '0%'

            if ws.cell(id_row + 21, mar_col).value >= 0.05:
                ws.cell(id_row + 21, tar_col).font = font_blue
                ws.cell(id_row + 21, mar_col).font = font_blue
                ws.cell(id_row + 21, res_col).value = 'PASS'
                ws.cell(id_row + 21, res_col).fill = fill_green

            if 0 < ws.cell(id_row + 21, mar_col).value < 0.05:
                ws.cell(id_row + 21, tar_col).fill = fill_yellow
                ws.cell(id_row + 21, mar_col).fill = fill_yellow
                ws.cell(id_row + 21, res_col).value = 'Marginal Pass'
                ws.cell(id_row + 21, res_col).fill = fill_yellow
        else:
            ws.cell(id_row + 21, tar_col).fill = fill_red
            ws.cell(id_row + 21, res_col).value = 'FAIL'
            ws.cell(id_row + 21, res_col).fill = fill_red

        # Thd margin
        if ws.cell(id_row + 22, tar_col).value > ws.cell(id_row + 22, tar_col - 2).value:
            ws.cell(id_row + 22, tar_col).font = font_blue
            if ws.cell(id_row + 22, tar_col - 2).value == 300:
                ws.cell(id_row + 22, mar_col).value = (ws.cell(id_row + 22, tar_col).value - 300)/300
                ws.cell(id_row + 22, mar_col).number_format = '0%'

                if ws.cell(id_row + 22, mar_col).value >= 0.05:
                    ws.cell(id_row + 22, tar_col).font = font_blue
                    ws.cell(id_row + 22, mar_col).font = font_blue
                    ws.cell(id_row + 22, res_col).value = 'PASS'
                    ws.cell(id_row + 22, res_col).fill = fill_green

                if 0 < ws.cell(id_row + 22, mar_col).value < 0.05:
                    ws.cell(id_row + 22, tar_col).fill = fill_yellow
                    ws.cell(id_row + 22, mar_col).fill = fill_yellow
                    ws.cell(id_row + 22, res_col).value = 'Marginal Pass'
                    ws.cell(id_row + 22, res_col).fill = fill_yellow

            else:
                ws.cell(id_row + 22, mar_col).value = '/'
                ws.cell(id_row + 22, mar_col).font = font_blue
                ws.cell(id_row + 22, res_col).value = 'PASS'
                ws.cell(id_row + 22, res_col).fill = fill_green

        else:
            ws.cell(id_row + 22, tar_col).fill = fill_red
            ws.cell(id_row + 22, res_col).value = 'FAIL'
            ws.cell(id_row + 22, res_col).fill = fill_red

        # Sta margin
        if ws.cell(id_row + 23, tar_col).value > ws.cell(id_row + 23, tar_col-2).value:
            ws.cell(id_row+23, mar_col).value = \
                (ws.cell(id_row + 23, tar_col).value - ws.cell(id_row + 23, tar_col-2).value) / \
                ws.cell(id_row + 23, tar_col-2).value

            ws.cell(id_row + 23, mar_col).number_format = '0%'

            if ws.cell(id_row + 23, mar_col).value >= 0.05:
                ws.cell(id_row + 23, tar_col).font = font_blue
                ws.cell(id_row + 23, mar_col).font = font_blue
                ws.cell(id_row + 23, res_col).value = 'PASS'
                ws.cell(id_row + 23, res_col).fill = fill_green

            if 0 < ws.cell(id_row + 23, mar_col).value < 0.05:
                ws.cell(id_row + 23, tar_col).fill = fill_yellow
                ws.cell(id_row + 23, mar_col).fill = fill_yellow
                ws.cell(id_row + 23, res_col).value = 'Marginal Pass'
                ws.cell(id_row + 23, res_col).fill = fill_yellow
        else:
            ws.cell(id_row + 23, tar_col).fill = fill_red
            ws.cell(id_row + 23, res_col).value = 'FAIL'
            ws.cell(id_row + 23, res_col).fill = fill_red

        # Sto margin
        if ws.cell(id_row + 24, tar_col).value > ws.cell(id_row + 24, tar_col-2).value:
            ws.cell(id_row+24, mar_col).value = \
                (ws.cell(id_row + 24, tar_col).value - ws.cell(id_row + 24, tar_col-2).value) / \
                ws.cell(id_row + 24, tar_col-2).value

            ws.cell(id_row + 24, mar_col).number_format = '0%'

            if ws.cell(id_row + 24, mar_col).value >= 0.05:
                ws.cell(id_row + 24, tar_col).font = font_blue
                ws.cell(id_row + 24, mar_col).font = font_blue
                ws.cell(id_row + 24, res_col).value = 'PASS'
                ws.cell(id_row + 24, res_col).fill = fill_green

            if 0 < ws.cell(id_row + 24, mar_col).value < 0.05:
                ws.cell(id_row + 24, tar_col).fill = fill_yellow
                ws.cell(id_row + 24, mar_col).fill = fill_yellow
                ws.cell(id_row + 24, res_col).value = 'Marginal Pass'
                ws.cell(id_row + 24, res_col).fill = fill_yellow
        else:
            ws.cell(id_row + 24, tar_col).fill = fill_red
            ws.cell(id_row + 24, res_col).value = 'FAIL'
            ws.cell(id_row + 24, res_col).fill = fill_red

        # SCL Monotonic
        ws.cell(id_row + 25, tar_col).font = font_blue
        ws.cell(id_row+25, mar_col).value = '/'
        ws.cell(id_row + 25, mar_col).font = font_blue
        ws.cell(id_row + 25, res_col).value = 'PASS'
        ws.cell(id_row + 25, res_col).fill = fill_green

        # wb.save('13 R0A I2C test report_v1.0.xlsx')

        # insert captures
        id_capture = os.path.join(cap_path, idn)

        # insert freq/high time/low time
        freq_path = id_capture + '\\freq.png'
        # freq_img = Image(freq_path)
        # freq_img.anchor = 'B' + str(id_row+28)
        freq_img_pos = 'B' + str(id_row + 28)
        # ws.add_image(freq_img)
        inserImg(ws, freq_path, freq_img_pos)

        # insert voltage
        vol_path = id_capture + '\\v.png'
        # vol_img = Image(vol_path)
        vol_img_pos = 'E' + str(id_row + 28)
        inserImg(ws, vol_path, vol_img_pos)

        # insert rise/fall
        rf_path = id_capture + '\\rf.png'
        # vol_img = Image(vol_path)
        rf_img_pos = 'B' + str(id_row + 28 + 11)
        inserImg(ws, rf_path, rf_img_pos)

        # insert SDA setup
        tsu_path = id_capture + '\\tsu.png'
        tsu_img_pos = 'E' + str(id_row + 28 + 11)
        inserImg(ws, tsu_path, tsu_img_pos)

        # insert thd
        thd_path = id_capture + '\\thd.png'
        thd_img_pos = 'B' + str(id_row + 28 + 11 + 11)
        inserImg(ws, thd_path, thd_img_pos)

        # insert sta
        sta_path = id_capture + '\\sta.png'
        sta_img_pos = 'E' + str(id_row + 28 + 11 + 11)
        inserImg(ws, sta_path, sta_img_pos)

        # insert sto
        sto_path = id_capture + '\\sto.png'
        sto_img_pos = 'B' + str(id_row + 28 + 11 + 11 + 11)
        inserImg(ws, sto_path, sto_img_pos)

        # insert SCL rising edge
        r_path = id_capture + '\\r.png'
        r_img_pos = 'E' + str(id_row + 28 + 11 + 11 + 11)
        inserImg(ws, r_path, r_img_pos)

        # insert SCL falling edge
        f_path = id_capture + '\\f.png'
        f_img_pos = 'B' + str(id_row + 28 + 11 + 11 + 11 + 11)
        inserImg(ws, f_path, f_img_pos)

        # wb.save('13 R0A I2C test report_v1.0.xlsx')

    # margin calculation
    if int(num_csv) == 1:

        # SCL_MAX margin
        if type(ws.cell(id_row + 6, tar_col - 1).value) is str:
            if ws.cell(id_row + 6, tar_col - 1).value == 'NA' or '/':
                # print(type(ws.cell(id_row + 6, tar_col - 1).value))
                ws.cell(id_row + 6, tar_col).font = font_blue
                ws.cell(id_row + 6, mar_col).value = '/'
                ws.cell(id_row + 6, res_col).value = 'For Reference'

        else:
            if ws.cell(id_row + 6, tar_col).value < ws.cell(id_row + 6, tar_col - 1).value:

                ws.cell(id_row + 6, mar_col).value = \
                    (ws.cell(id_row + 6, tar_col - 1).value - ws.cell(id_row + 6, tar_col).value) / \
                    ws.cell(id_row + 6, tar_col - 1).value
                ws.cell(id_row + 6, mar_col).number_format = '0%'

                if ws.cell(id_row + 6, mar_col).value >= 0.05:
                    ws.cell(id_row + 6, tar_col).font = font_blue
                    ws.cell(id_row + 6, mar_col).font = font_blue
                    ws.cell(id_row + 6, res_col).value = 'PASS'
                    ws.cell(id_row + 6, res_col).fill = fill_green

                if 0 < ws.cell(id_row + 6, mar_col).value < 0.05:
                    ws.cell(id_row + 6, tar_col).fill = fill_yellow
                    ws.cell(id_row + 6, mar_col).fill = fill_yellow
                    ws.cell(id_row + 6, res_col).value = 'Marginal Pass'
                    ws.cell(id_row + 6, res_col).fill = fill_yellow
            else:
                ws.cell(id_row + 6, tar_col).fill = fill_red
                # ws.cell(id_row + 9, mar_col).fill = fill_yellow
                ws.cell(id_row + 6, res_col).value = 'FAIL'
                ws.cell(id_row + 6, res_col).fill = fill_red

        # SCL_Min margin
        if type(ws.cell(id_row + 7, tar_col - 2).value) is str:
            if ws.cell(id_row + 7, tar_col - 2).value == 'NA' or '/':
                ws.cell(id_row + 7, tar_col).font = font_blue
                ws.cell(id_row + 7, mar_col).value = '/'
                ws.cell(id_row + 7, res_col).value = 'For Reference'

        else:
            if ws.cell(id_row + 7, tar_col).value > ws.cell(id_row + 7, tar_col - 2).value:

                ws.cell(id_row + 7, mar_col).value = \
                    (ws.cell(id_row + 7, tar_col - 2).value - ws.cell(id_row + 7, tar_col).value) / \
                    ws.cell(id_row + 7, tar_col - 2).value
                ws.cell(id_row + 7, mar_col).number_format = '0%'

                if ws.cell(id_row + 7, mar_col).value >= 0.05:
                    ws.cell(id_row + 7, tar_col).font = font_blue
                    ws.cell(id_row + 7, mar_col).font = font_blue
                    ws.cell(id_row + 7, res_col).value = 'PASS'
                    ws.cell(id_row + 7, res_col).fill = fill_green

                if 0 < ws.cell(id_row + 7, mar_col).value < 0.05:
                    ws.cell(id_row + 7, tar_col).fill = fill_yellow
                    ws.cell(id_row + 7, mar_col).fill = fill_yellow
                    ws.cell(id_row + 7, res_col).value = 'Marginal Pass'
                    ws.cell(id_row + 7, res_col).fill = fill_yellow
            else:
                ws.cell(id_row + 7, tar_col).fill = fill_red
                # ws.cell(id_row + 9, mar_col).fill = fill_yellow
                ws.cell(id_row + 7, res_col).value = 'FAIL'
                ws.cell(id_row + 7, res_col).fill = fill_red

        # SCL_H margin
        if type(ws.cell(id_row + 8, tar_col-2).value) is str:
            if ws.cell(id_row + 8, tar_col-2).value == 'NA' or '/':
                ws.cell(id_row + 8, tar_col).font = font_blue
                ws.cell(id_row + 8, mar_col).value = '/'
                ws.cell(id_row + 8, mar_col).font = font_blue
                ws.cell(id_row + 8, res_col).value = 'For Reference'

        elif ws.cell(id_row + 8, tar_col).value > ws.cell(id_row + 8, tar_col - 2).value:
            ws.cell(id_row + 8, mar_col).value = \
                (ws.cell(id_row + 8, tar_col).value - ws.cell(id_row + 8, tar_col - 2).value) / \
                ws.cell(id_row + 8, tar_col - 2).value
            ws.cell(id_row + 8, mar_col).number_format = '0%'

            if ws.cell(id_row + 8, mar_col).value >= 0.05:
                ws.cell(id_row + 8, tar_col).font = font_blue
                ws.cell(id_row + 8, mar_col).font = font_blue
                ws.cell(id_row + 8, res_col).value = 'PASS'
                ws.cell(id_row + 8, res_col).fill = fill_green

            if 0 < ws.cell(id_row + 8, mar_col).value < 0.05:
                ws.cell(id_row + 8, tar_col).fill = fill_yellow
                ws.cell(id_row + 8, mar_col).fill = fill_yellow
                ws.cell(id_row + 8, res_col).value = 'Marginal Pass'
                ws.cell(id_row + 8, res_col).fill = fill_yellow
        else:
            ws.cell(id_row + 8, tar_col).fill = fill_red
            ws.cell(id_row + 8, res_col).value = 'FAIL'
            ws.cell(id_row + 8, res_col).fill = fill_red

        # SCL_L margin
        if type(ws.cell(id_row + 9, tar_col-1).value) is str:
            if ws.cell(id_row + 9, tar_col-1).value == 'NA' or '/':
                ws.cell(id_row + 9, tar_col).font = font_blue
                ws.cell(id_row + 9, mar_col).value = '/'
                ws.cell(id_row + 9, mar_col).font = font_blue
                ws.cell(id_row + 9, res_col).value = 'For Reference'

        elif type(ws.cell(id_row + 9, tar_col-2).value) is float:
            if ws.cell(id_row + 9, tar_col-2).value < ws.cell(id_row + 9, tar_col).value < ws.cell(id_row + 9, tar_col - 1).value:
                if ws.cell(id_row + 9, tar_col-2).value == 0.45:
                    min_SCL_Vil = ws.cell(id_row + 9, tar_col).value - ws.cell(id_row + 9, tar_col - 2).value
                    max_SCL_Vil = ws.cell(id_row + 9, tar_col - 1).value - ws.cell(id_row + 9, tar_col).value
                    if min_SCL_Vil < max_SCL_Vil:
                        ws.cell(id_row + 9, mar_col).value = \
                            min_SCL_Vil / (ws.cell(id_row + 9, tar_col-1).value - ws.cell(id_row + 9, tar_col - 2).value)
                    else:
                        ws.cell(id_row + 9, mar_col).value = \
                            max_SCL_Vil / (ws.cell(id_row + 9, tar_col - 1).value - ws.cell(id_row + 9, tar_col - 2).value)
                else:
                    ws.cell(id_row + 9, mar_col).value = \
                        (ws.cell(id_row + 9, tar_col - 1).value - ws.cell(id_row + 9, tar_col).value) / \
                        ws.cell(id_row + 9, tar_col - 1).value
                ws.cell(id_row + 9, mar_col).number_format = '0%'

                if ws.cell(id_row + 9, mar_col).value >= 0.05:
                    ws.cell(id_row + 9, tar_col).font = font_blue
                    ws.cell(id_row + 9, mar_col).font = font_blue
                    ws.cell(id_row + 9, res_col).value = 'PASS'
                    ws.cell(id_row + 9, res_col).fill = fill_green

                if 0 < ws.cell(id_row + 9, mar_col).value < 0.05:
                    ws.cell(id_row + 9, tar_col).fill = fill_yellow
                    ws.cell(id_row + 9, mar_col).fill = fill_yellow
                    ws.cell(id_row + 9, res_col).value = 'Marginal Pass'
                    ws.cell(id_row + 9, res_col).fill = fill_yellow

            else:
                ws.cell(id_row + 9, tar_col).fill = fill_red
                ws.cell(id_row + 9, res_col).value = 'FAIL'
                ws.cell(id_row + 9, res_col).fill = fill_red

        else:
            if ws.cell(id_row + 9, tar_col).value < ws.cell(id_row + 9, tar_col - 1).value:
                ws.cell(id_row + 9, mar_col).value = \
                    (ws.cell(id_row + 9, tar_col - 1).value - ws.cell(id_row + 9, tar_col).value) / \
                    ws.cell(id_row + 9, tar_col - 1).value
                ws.cell(id_row + 9, mar_col).number_format = '0%'

                if ws.cell(id_row + 9, mar_col).value >= 0.05:
                    ws.cell(id_row + 9, tar_col).font = font_blue
                    ws.cell(id_row + 9, mar_col).font = font_blue
                    ws.cell(id_row + 9, res_col).value = 'PASS'
                    ws.cell(id_row + 9, res_col).fill = fill_green

                if 0 < ws.cell(id_row + 9, mar_col).value < 0.05:
                    ws.cell(id_row + 9, tar_col).fill = fill_yellow
                    ws.cell(id_row + 9, mar_col).fill = fill_yellow
                    ws.cell(id_row + 9, res_col).value = 'Marginal Pass'
                    ws.cell(id_row + 9, res_col).fill = fill_yellow
            else:
                ws.cell(id_row + 9, tar_col).fill = fill_red
                ws.cell(id_row + 9, res_col).value = 'FAIL'
                ws.cell(id_row + 9, res_col).fill = fill_red

        # SDA_MAX margin
        if type(ws.cell(id_row + 10, tar_col - 1).value) is str:
            if ws.cell(id_row + 10, tar_col - 1).value == 'NA' or '/':
                ws.cell(id_row + 10, tar_col).font = font_blue
                ws.cell(id_row + 10, mar_col).value = '/'
                ws.cell(id_row + 10, res_col).value = 'For Reference'

        else:
            if ws.cell(id_row + 10, tar_col).value < ws.cell(id_row + 10, tar_col - 1).value:

                ws.cell(id_row + 10, mar_col).value = \
                    (ws.cell(id_row + 10, tar_col - 1).value - ws.cell(id_row + 10, tar_col).value) / \
                    ws.cell(id_row + 10, tar_col - 1).value
                ws.cell(id_row + 10, mar_col).number_format = '0%'

                if ws.cell(id_row + 10, mar_col).value >= 0.05:
                    ws.cell(id_row + 10, tar_col).font = font_blue
                    ws.cell(id_row + 10, mar_col).font = font_blue
                    ws.cell(id_row + 10, res_col).value = 'PASS'
                    ws.cell(id_row + 10, res_col).fill = fill_green

                if 0 < ws.cell(id_row + 10, mar_col).value < 0.05:
                    ws.cell(id_row + 10, tar_col).fill = fill_yellow
                    ws.cell(id_row + 10, mar_col).fill = fill_yellow
                    ws.cell(id_row + 10, res_col).value = 'Marginal Pass'
                    ws.cell(id_row + 10, res_col).fill = fill_yellow
            else:
                ws.cell(id_row + 10, tar_col).fill = fill_red
                # ws.cell(id_row + 9, mar_col).fill = fill_yellow
                ws.cell(id_row + 10, res_col).value = 'FAIL'
                ws.cell(id_row + 10, res_col).fill = fill_red

        # SDA_Min margin
        if type(ws.cell(id_row + 11, tar_col - 2).value) is str:
            if ws.cell(id_row + 11, tar_col - 2).value == 'NA' or '/':
                ws.cell(id_row + 11, tar_col).font = font_blue
                ws.cell(id_row + 11, mar_col).value = '/'
                ws.cell(id_row + 11, res_col).value = 'For Reference'

        else:
            if ws.cell(id_row + 11, tar_col).value > ws.cell(id_row + 11, tar_col - 2).value:

                ws.cell(id_row + 11, mar_col).value = \
                    (ws.cell(id_row + 11, tar_col - 2).value - ws.cell(id_row + 11, tar_col).value) / \
                    ws.cell(id_row + 11, tar_col - 2).value
                ws.cell(id_row + 11, mar_col).number_format = '0%'

                if ws.cell(id_row + 11, mar_col).value >= 0.05:
                    ws.cell(id_row + 11, tar_col).font = font_blue
                    ws.cell(id_row + 11, mar_col).font = font_blue
                    ws.cell(id_row + 11, res_col).value = 'PASS'
                    ws.cell(id_row + 11, res_col).fill = fill_green

                if 0 < ws.cell(id_row + 11, mar_col).value < 0.05:
                    ws.cell(id_row + 11, tar_col).fill = fill_yellow
                    ws.cell(id_row + 11, mar_col).fill = fill_yellow
                    ws.cell(id_row + 11, res_col).value = 'Marginal Pass'
                    ws.cell(id_row + 11, res_col).fill = fill_yellow
            else:
                ws.cell(id_row + 11, tar_col).fill = fill_red
                # ws.cell(id_row + 9, mar_col).fill = fill_yellow
                ws.cell(id_row + 11, res_col).value = 'FAIL'
                ws.cell(id_row + 11, res_col).fill = fill_red

        # SDA_H margin
        if type(ws.cell(id_row + 12, tar_col - 2).value) is str:
            if ws.cell(id_row + 12, tar_col - 2).value == 'NA' or '/':
                ws.cell(id_row + 12, tar_col).font = font_blue
                ws.cell(id_row + 12, mar_col).value = '/'
                ws.cell(id_row + 12, mar_col).font = font_blue
                ws.cell(id_row + 12, res_col).value = 'For Reference'

        elif ws.cell(id_row + 12, tar_col).value > ws.cell(id_row + 12, tar_col - 2).value:
            ws.cell(id_row + 12, mar_col).value = \
                (ws.cell(id_row + 12, tar_col).value - ws.cell(id_row + 12, tar_col - 2).value) / \
                ws.cell(id_row + 12, tar_col - 2).value
            ws.cell(id_row + 12, mar_col).number_format = '0%'

            if ws.cell(id_row + 12, mar_col).value >= 0.05:
                ws.cell(id_row + 12, tar_col).font = font_blue
                ws.cell(id_row + 12, mar_col).font = font_blue
                ws.cell(id_row + 12, res_col).value = 'PASS'
                ws.cell(id_row + 12, res_col).fill = fill_green

            if 0 < ws.cell(id_row + 12, mar_col).value < 0.05:
                ws.cell(id_row + 12, tar_col).fill = fill_yellow
                ws.cell(id_row + 12, mar_col).fill = fill_yellow
                ws.cell(id_row + 12, res_col).value = 'Marginal Pass'
                ws.cell(id_row + 12, res_col).fill = fill_yellow
        else:
            ws.cell(id_row + 12, tar_col).fill = fill_red
            ws.cell(id_row + 12, res_col).value = 'FAIL'
            ws.cell(id_row + 12, res_col).fill = fill_red

        # SDA_L margin
        if type(ws.cell(id_row + 13, tar_col - 1).value) is str:
            if ws.cell(id_row + 13, tar_col - 1).value == 'NA' or '/':
                ws.cell(id_row + 13, tar_col).font = font_blue
                ws.cell(id_row + 13, mar_col).value = '/'
                ws.cell(id_row + 13, mar_col).font = font_blue
                ws.cell(id_row + 13, res_col).value = 'For Reference'

        elif ws.cell(id_row + 13, tar_col).value < ws.cell(id_row + 13, tar_col - 1).value:
            ws.cell(id_row + 13, mar_col).value = \
                (ws.cell(id_row + 13, tar_col - 1).value - ws.cell(id_row + 13, tar_col).value) / \
                ws.cell(id_row + 13, tar_col - 1).value
            ws.cell(id_row + 13, mar_col).number_format = '0%'

            if ws.cell(id_row + 13, mar_col).value >= 0.05:
                ws.cell(id_row + 13, tar_col).font = font_blue
                ws.cell(id_row + 13, mar_col).font = font_blue
                ws.cell(id_row + 13, res_col).value = 'PASS'
                ws.cell(id_row + 13, res_col).fill = fill_green

            if 0 < ws.cell(id_row + 13, mar_col).value < 0.05:
                ws.cell(id_row + 13, tar_col).fill = fill_yellow
                ws.cell(id_row + 13, mar_col).fill = fill_yellow
                ws.cell(id_row + 13, res_col).value = 'Marginal Pass'
                ws.cell(id_row + 13, res_col).fill = fill_yellow
        else:
            ws.cell(id_row + 13, tar_col).fill = fill_red
            ws.cell(id_row + 13, res_col).value = 'FAIL'
            ws.cell(id_row + 13, res_col).fill = fill_red

        # wb.save('13 R0A I2C test report_v1.0.xlsx')

        # insert captures
        id_capture = os.path.join(cap_path, idn)
        # insert voltage
        vol_path = id_capture + '\\v.png'
        # vol_img = Image(vol_path)
        vol_img_pos = 'B' + str(id_row + 16)
        inserImg(ws, vol_path, vol_img_pos)

        # wb.save('13 R0A I2C test report_v1.0.xlsx')

    # margin calculation
    if int(num_csv) == 6:
        # SDA_Max margin
        if type(ws.cell(id_row + 6, tar_col - 1).value) is str:
            if ws.cell(id_row + 6, tar_col - 1).value == 'NA' or '/':

                ws.cell(id_row + 6, tar_col).font = font_blue
                ws.cell(id_row + 6, mar_col).value = '/'
                ws.cell(id_row + 6, res_col).value = 'For Reference'

        else:
            if ws.cell(id_row + 6, tar_col).value < ws.cell(id_row + 6, tar_col - 1).value:

                ws.cell(id_row + 6, mar_col).value = \
                    (ws.cell(id_row + 6, tar_col - 1).value - ws.cell(id_row + 6, tar_col).value) / \
                    ws.cell(id_row + 6, tar_col - 1).value
                ws.cell(id_row + 6, mar_col).number_format = '0%'

                if ws.cell(id_row + 6, mar_col).value >= 0.05:
                    ws.cell(id_row + 6, tar_col).font = font_blue
                    ws.cell(id_row + 6, mar_col).font = font_blue
                    ws.cell(id_row + 6, res_col).value = 'PASS'
                    ws.cell(id_row + 6, res_col).fill = fill_green

                if 0 < ws.cell(id_row + 6, mar_col).value < 0.05:
                    ws.cell(id_row + 6, tar_col).fill = fill_yellow
                    ws.cell(id_row + 6, mar_col).fill = fill_yellow
                    ws.cell(id_row + 6, res_col).value = 'Marginal Pass'
                    ws.cell(id_row + 6, res_col).fill = fill_yellow
            else:
                ws.cell(id_row + 6, tar_col).fill = fill_red
                # ws.cell(id_row + 9, mar_col).fill = fill_yellow
                ws.cell(id_row + 6, res_col).value = 'FAIL'
                ws.cell(id_row + 6, res_col).fill = fill_red

        # SDA_Min margin
        if type(ws.cell(id_row + 7, tar_col - 2).value) is str:
            if ws.cell(id_row + 7, tar_col - 2).value == 'NA' or '/':
                ws.cell(id_row + 7, tar_col).font = font_blue
                ws.cell(id_row + 7, mar_col).value = '/'
                ws.cell(id_row + 7, res_col).value = 'For Reference'

        else:
            if ws.cell(id_row + 7, tar_col).value > ws.cell(id_row + 7, tar_col - 2).value:

                ws.cell(id_row + 7, mar_col).value = \
                    (ws.cell(id_row + 7, tar_col - 2).value - ws.cell(id_row + 7, tar_col).value) / \
                    ws.cell(id_row + 7, tar_col - 2).value
                ws.cell(id_row + 7, mar_col).number_format = '0%'

                if ws.cell(id_row + 7, mar_col).value >= 0.05:
                    ws.cell(id_row + 7, tar_col).font = font_blue
                    ws.cell(id_row + 7, mar_col).font = font_blue
                    ws.cell(id_row + 7, res_col).value = 'PASS'
                    ws.cell(id_row + 7, res_col).fill = fill_green

                if 0 < ws.cell(id_row + 7, mar_col).value < 0.05:
                    ws.cell(id_row + 7, tar_col).fill = fill_yellow
                    ws.cell(id_row + 7, mar_col).fill = fill_yellow
                    ws.cell(id_row + 7, res_col).value = 'Marginal Pass'
                    ws.cell(id_row + 7, res_col).fill = fill_yellow
            else:
                ws.cell(id_row + 7, tar_col).fill = fill_red
                # ws.cell(id_row + 9, mar_col).fill = fill_yellow
                ws.cell(id_row + 7, res_col).value = 'FAIL'
                ws.cell(id_row + 7, res_col).fill = fill_red

        # SDA_H margin
        if ws.cell(id_row + 8, tar_col).value > ws.cell(id_row + 8, tar_col - 2).value:
            ws.cell(id_row + 8, mar_col).value = \
                (ws.cell(id_row + 8, tar_col).value - ws.cell(id_row + 8, tar_col - 2).value) / \
                ws.cell(id_row + 8, tar_col - 2).value
            ws.cell(id_row + 8, mar_col).number_format = '0%'

            if ws.cell(id_row + 8, mar_col).value >= 0.05:
                ws.cell(id_row + 8, tar_col).font = font_blue
                ws.cell(id_row + 8, mar_col).font = font_blue
                ws.cell(id_row + 8, res_col).value = 'PASS'
                ws.cell(id_row + 8, res_col).fill = fill_green

            if 0 < ws.cell(id_row + 8, mar_col).value < 0.05:
                ws.cell(id_row + 8, tar_col).fill = fill_yellow
                ws.cell(id_row + 8, mar_col).fill = fill_yellow
                ws.cell(id_row + 8, res_col).value = 'Marginal Pass'
                ws.cell(id_row + 8, res_col).fill = fill_yellow
        else:
            ws.cell(id_row + 8, tar_col).fill = fill_red
            ws.cell(id_row + 8, res_col).value = 'FAIL'
            ws.cell(id_row + 8, res_col).fill = fill_red

        # SDA_L margin
        if ws.cell(id_row + 9, tar_col).value < ws.cell(id_row + 9, tar_col - 1).value:
            ws.cell(id_row + 9, mar_col).value = \
                (ws.cell(id_row + 9, tar_col - 1).value - ws.cell(id_row + 9, tar_col).value) / \
                ws.cell(id_row + 9, tar_col - 1).value
            ws.cell(id_row + 9, mar_col).number_format = '0%'

            if ws.cell(id_row + 9, mar_col).value >= 0.05:
                ws.cell(id_row + 9, tar_col).font = font_blue
                ws.cell(id_row + 9, mar_col).font = font_blue
                ws.cell(id_row + 9, res_col).value = 'PASS'
                ws.cell(id_row + 9, res_col).fill = fill_green

            if 0 < ws.cell(id_row + 9, mar_col).value < 0.05:
                ws.cell(id_row + 9, tar_col).fill = fill_yellow
                ws.cell(id_row + 9, mar_col).fill = fill_yellow
                ws.cell(id_row + 9, res_col).value = 'Marginal Pass'
                ws.cell(id_row + 9, res_col).fill = fill_yellow
        else:
            ws.cell(id_row + 9, tar_col).fill = fill_red
            ws.cell(id_row + 9, res_col).value = 'FAIL'
            ws.cell(id_row + 9, res_col).fill = fill_red

        # SDA_R margin
        if ws.cell(id_row+10, tar_col-2).value < ws.cell(id_row+10, tar_col).value < ws.cell(id_row+10, tar_col-1).value:
            min_SCL_R = ws.cell(id_row+10, tar_col).value - ws.cell(id_row+10, tar_col-2).value
            max_SCL_R = ws.cell(id_row+10, tar_col-1).value - ws.cell(id_row+10, tar_col).value
            if max_SCL_R < min_SCL_R:
                ws.cell(id_row + 10, mar_col).value = max_SCL_R / ws.cell(id_row + 10, tar_col - 1).value
            else:
                if ws.cell(id_row+10, tar_col-2).value == 0:
                    ws.cell(id_row + 10, mar_col).value = min_SCL_R / ws.cell(id_row + 10, tar_col - 1).value
                else:
                    ws.cell(id_row + 10, mar_col).value = min_SCL_R / ws.cell(id_row + 10, tar_col - 2).value

            ws.cell(id_row + 10, mar_col).number_format = '0%'

            if ws.cell(id_row + 10, mar_col).value >= 0.05:
                ws.cell(id_row + 10, tar_col).font = font_blue
                ws.cell(id_row + 10, mar_col).font = font_blue
                ws.cell(id_row + 10, res_col).value = 'PASS'
                ws.cell(id_row + 10, res_col).fill = fill_green

            if 0 < ws.cell(id_row + 10, mar_col).value < 0.05:
                ws.cell(id_row + 10, tar_col).fill = fill_yellow
                ws.cell(id_row + 10, mar_col).fill = fill_yellow
                ws.cell(id_row + 10, res_col).value = 'Marginal Pass'
                ws.cell(id_row + 10, res_col).fill = fill_yellow
        else:
            ws.cell(id_row + 10, tar_col).fill = fill_red
            ws.cell(id_row + 10, res_col).value = 'FAIL'
            ws.cell(id_row + 10, res_col).fill = fill_red

        # SDA_Fall margin
        if ws.cell(id_row+11, tar_col-2).value < ws.cell(id_row+11, tar_col).value < ws.cell(id_row+11, tar_col-1).value:
            ws.cell(id_row+11, mar_col).value = 1 - ws.cell(id_row+11, tar_col).value / ws.cell(id_row+11, tar_col-1).value

            ws.cell(id_row + 11, mar_col).number_format = '0%'

            if ws.cell(id_row + 11, mar_col).value >= 0.05:
                ws.cell(id_row + 11, tar_col).font = font_blue
                ws.cell(id_row + 11, mar_col).font = font_blue
                ws.cell(id_row + 11, res_col).value = 'PASS'
                ws.cell(id_row + 11, res_col).fill = fill_green

            if 0 < ws.cell(id_row + 11, mar_col).value < 0.05:
                ws.cell(id_row + 11, tar_col).fill = fill_yellow
                ws.cell(id_row + 11, mar_col).fill = fill_yellow
                ws.cell(id_row + 11, res_col).value = 'Marginal Pass'
                ws.cell(id_row + 11, res_col).fill = fill_yellow
        else:
            ws.cell(id_row + 11, tar_col).fill = fill_red
            ws.cell(id_row + 11, res_col).value = 'FAIL'
            ws.cell(id_row + 11, res_col).fill = fill_red

        # Tsu margin
        if ws.cell(id_row + 12, tar_col).value > ws.cell(id_row + 12, tar_col-2).value:
            ws.cell(id_row+12, mar_col).value = \
                (ws.cell(id_row + 12, tar_col).value - ws.cell(id_row + 12, tar_col-2).value) / \
                ws.cell(id_row + 12, tar_col-2).value

            ws.cell(id_row + 12, mar_col).number_format = '0%'

            if ws.cell(id_row + 12, mar_col).value >= 0.05:
                ws.cell(id_row + 12, tar_col).font = font_blue
                ws.cell(id_row + 12, mar_col).font = font_blue
                ws.cell(id_row + 12, res_col).value = 'PASS'
                ws.cell(id_row + 12, res_col).fill = fill_green

            if 0 < ws.cell(id_row + 12, mar_col).value < 0.05:
                ws.cell(id_row + 12, tar_col).fill = fill_yellow
                ws.cell(id_row + 12, mar_col).fill = fill_yellow
                ws.cell(id_row + 12, res_col).value = 'Marginal Pass'
                ws.cell(id_row + 12, res_col).fill = fill_yellow
        else:
            ws.cell(id_row + 12, tar_col).fill = fill_red
            ws.cell(id_row + 12, res_col).value = 'FAIL'
            ws.cell(id_row + 12, res_col).fill = fill_red

        # Thd margin
        if ws.cell(id_row + 13, tar_col).value > ws.cell(id_row + 13, tar_col - 2).value:
            ws.cell(id_row + 13, tar_col).font = font_blue
            ws.cell(id_row + 13, mar_col).value = '/'
            ws.cell(id_row + 13, mar_col).font = font_blue
            ws.cell(id_row + 13, res_col).value = 'PASS'
            ws.cell(id_row + 13, res_col).fill = fill_green

        else:
            ws.cell(id_row + 13, tar_col).fill = fill_red
            ws.cell(id_row + 13, res_col).value = 'FAIL'
            ws.cell(id_row + 13, res_col).fill = fill_red

        # Sta margin
        if ws.cell(id_row + 14, tar_col).value > ws.cell(id_row + 14, tar_col-2).value:
            ws.cell(id_row+14, mar_col).value = \
                (ws.cell(id_row + 14, tar_col).value - ws.cell(id_row + 14, tar_col-2).value) / \
                ws.cell(id_row + 14, tar_col-2).value

            ws.cell(id_row + 14, mar_col).number_format = '0%'

            if ws.cell(id_row + 14, mar_col).value >= 0.05:
                ws.cell(id_row + 14, tar_col).font = font_blue
                ws.cell(id_row + 14, mar_col).font = font_blue
                ws.cell(id_row + 14, res_col).value = 'PASS'
                ws.cell(id_row + 14, res_col).fill = fill_green

            if 0 < ws.cell(id_row + 14, mar_col).value < 0.05:
                ws.cell(id_row + 14, tar_col).fill = fill_yellow
                ws.cell(id_row + 14, mar_col).fill = fill_yellow
                ws.cell(id_row + 14, res_col).value = 'Marginal Pass'
                ws.cell(id_row + 14, res_col).fill = fill_yellow
        else:
            ws.cell(id_row + 14, tar_col).fill = fill_red
            ws.cell(id_row + 14, res_col).value = 'FAIL'
            ws.cell(id_row + 14, res_col).fill = fill_red

        # Sto margin
        if ws.cell(id_row + 15, tar_col).value > ws.cell(id_row + 15, tar_col-2).value:
            ws.cell(id_row+15, mar_col).value = \
                (ws.cell(id_row + 15, tar_col).value - ws.cell(id_row + 15, tar_col-2).value) / \
                ws.cell(id_row + 15, tar_col-2).value

            ws.cell(id_row + 15, mar_col).number_format = '0%'

            if ws.cell(id_row + 15, mar_col).value >= 0.05:
                ws.cell(id_row + 15, tar_col).font = font_blue
                ws.cell(id_row + 15, mar_col).font = font_blue
                ws.cell(id_row + 15, res_col).value = 'PASS'
                ws.cell(id_row + 15, res_col).fill = fill_green

            if 0 < ws.cell(id_row + 15, mar_col).value < 0.05:
                ws.cell(id_row + 15, tar_col).fill = fill_yellow
                ws.cell(id_row + 15, mar_col).fill = fill_yellow
                ws.cell(id_row + 15, res_col).value = 'Marginal Pass'
                ws.cell(id_row + 15, res_col).fill = fill_yellow
        else:
            ws.cell(id_row + 15, tar_col).fill = fill_red
            ws.cell(id_row + 15, res_col).value = 'FAIL'
            ws.cell(id_row + 15, res_col).fill = fill_red

        # wb.save('13 R0A I2C test plan_v1.0.xlsx')

        # insert captures
        id_capture = os.path.join(cap_path, idn)

        # insert voltage
        vol_path = id_capture + '\\v.png'
        # vol_img = Image(vol_path)
        vol_img_pos = 'B' + str(id_row + 18)
        inserImg(ws, vol_path, vol_img_pos)

        # insert rise/fall
        rf_path = id_capture + '\\rf.png'
        # vol_img = Image(vol_path)
        rf_img_pos = 'E' + str(id_row + 18)
        inserImg(ws, rf_path, rf_img_pos)

        # insert SDA setup
        tsu_path = id_capture + '\\tsu.png'
        tsu_img_pos = 'B' + str(id_row + 18 + 11)
        inserImg(ws, tsu_path, tsu_img_pos)

        # insert thd
        thd_path = id_capture + '\\thd.png'
        thd_img_pos = 'E' + str(id_row + 18 + 11)
        inserImg(ws, thd_path, thd_img_pos)

        # insert sta
        sta_path = id_capture + '\\sta.png'
        sta_img_pos = 'B' + str(id_row + 18 + 11 + 11)
        inserImg(ws, sta_path, sta_img_pos)

        # insert sto
        sto_path = id_capture + '\\sto.png'
        sto_img_pos = 'E' + str(id_row + 18 + 11 + 11)
        inserImg(ws, sto_path, sto_img_pos)

wb.save('1 [TPDR-05-1-001]R0B I2C test report_v0.3.xlsx')
