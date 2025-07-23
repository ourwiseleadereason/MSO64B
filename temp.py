
import os
import pandas as pd
import openpyxl

rootPath = os.getcwd()

planPath = os.path.join(rootPath, 'test plan')

resultPath = os.path.join(rootPath, 'test data')
pnPath = os.path.join(resultPath, 'Power noise')
item_list = os.listdir(pnPath)
print(item_list)


# scan the test data folder and get the test data path
def get_data_path(ipPath, item_id):

    idPath = os.path.join(ipPath, item_id)
    # print(os.listdir(idPath))
    dataName = os.listdir(idPath)
    # print(dataName)
    for data in dataName:
        if 'L_V.csv' in data:
            # print(data)
            L_V_path = os.path.join(idPath, data)
            # print(L_V_path)
        # else:
        #     L_V_path = None

        if 'H_V.csv' in data:
            # print(data)
            H_V_path = os.path.join(idPath, data)
            # print(H_V_path)
        # else:
        #     H_V_path = None

        if 'L_N.csv' in data:
            # print(data)
            L_N_path = os.path.join(idPath, data)
            # print(L_N_path)
        # else:
        #     L_N_path = None

        if 'H_N.csv' in data:
            # print(data)
            H_N_path = os.path.join(idPath, data)
            # print(H_N_path)
        # else:
        #     H_N_path = None

    return L_V_path, H_V_path, L_N_path, H_N_path


# handle csv files to get the measured data
def csv_handle(path, target_row, target_col):

    largest_column_count = 0
    # target_row = 'RMS'
    row_target = 0
    # target_col = 'Mean\''
    col_target = 0

    with open(path, 'r') as f:
        # reader = csv.reader(f)
        lines = f.readlines()
        # print(lines)
        for line in lines:
            column_count = len(line.split(',')) + 1
            largest_column_count = column_count if largest_column_count < column_count else largest_column_count
    f.close()

    column_names = [i for i in range(0, largest_column_count)]

    df = pd.read_csv(path, header=None, delimiter=',', names=column_names)
    row_count = df.shape[0]
    col_count = df.shape[1]
    # print(row_count, col_count)

    for row in range(0, int(row_count)):
        for col in range(0, int(col_count)):

            if df[col][row] == target_row:
                row_target = row
                # print(row_target)
            if df[col][row] == target_col:
                col_target = col
    # print(row_target, col_target)
    mea_data = df[col_target][row_target]
    # print(col_target, row_target)
    return mea_data


# scan the test plan and get the col of test id and the test items.
def get_report_col(file):

    # wb = openpyxl.load_workbook(r'E:\Grace\power noise data processing\11 Power noise&Voltage.xlsx')
    wb = openpyxl.load_workbook(file)

    sheet = wb['Power Noise']
    rows = sheet.max_row
    columns = sheet.max_column

    ID_col, L_V_col, H_V_col, L_N_col, H_N_col = 0, 0, 0, 0, 0

    for row in range(1, rows + 1):
        for column in range(1, columns + 1):
            value = sheet.cell(row, column).value

            if 'ID' in str(value):
                ID_col = column
                # print(ID_col)

            if 'Voltage at light' in str(value):
                L_V_col = column
                # print(L_V_col)

            if 'Voltage at full' in str(value):
                H_V_col = column
                # print(H_V_col)

            if 'Noise&Ripple at light load' in str(value):
                L_N_col = column
                # print(L_N_col)

            if 'Noise&Ripple at full load' in str(value):
                H_N_col = column
                # print(H_N_col)

    return ID_col, L_V_col, H_V_col, L_N_col, H_N_col


# scan the test plan and get the row of test id
def get_id_row(file, target_id):

    # wb = openpyxl.load_workbook(r'E:\Grace\Data Processing\test plan\11. Power Niose & Voltage.xlsx')
    global id_row_number
    wb = openpyxl.load_workbook(file)

    sheet = wb['Power Noise']
    # sheet = wb['Sheet1']
    rows = sheet.max_row
    # print(rows)
    columns = sheet.max_column
    # print(columns)

    for row in range(1, rows+1):
        for column in range(1, columns+1):
            value = sheet.cell(row, column).value

            if str(value) == target_id:
                id_row_number = row
                # print(id_row_number)
    return id_row_number


plan = r'E:\Grace\test plan\17 Power Noise.xlsx'
wb = openpyxl.load_workbook(plan)
# wb = openpyxl.load_workbook(file)

sheet = wb['Power Noise']

for id_number in item_list:

    print(id_number)
    row_count = get_id_row(plan, id_number)
    print(row_count)

    id_col, lv_col, hv_col, ln_col, hn_col = get_report_col(plan)
    print(id_col, lv_col, hv_col, ln_col, hn_col)
    # print(sheet.cell(4, 13).value)

    # sheet.cell(row=6, column=10).value = data
    lvPath, hvPath, lnPath, hnPath = get_data_path(pnPath, id_number)
    print(lvPath)
    lv_value = csv_handle(lvPath, ' RMS', 'Mean\'')
    print(row_count,lv_col,lv_value)
    sheet.cell(row_count, lv_col).value = lv_value

    hv_value = csv_handle(hvPath, ' RMS', 'Mean\'')
    print(row_count, hv_col, hv_value)
    sheet.cell(row_count, hv_col).value = hv_value

    ln_value = csv_handle(lnPath, 'Peak-to-Peak', 'Max\'')
    print(row_count, ln_col, ln_value)
    sheet.cell(row_count, ln_col).value = ln_value

    hn_value = csv_handle(hnPath, 'Peak-to-Peak', 'Max\'')
    print(row_count, hn_col, hn_value)
    sheet.cell(row_count, hn_col).value = hn_value

wb.save(plan)

