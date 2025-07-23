import os
import openpyxl
import csv
import pandas as pd
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import colors
from openpyxl.styles import Alignment
from openpyxl.drawing.image import Image

#
# # handle csv files to get the measured data
# def csv_handle(path, target_row, target_col):
#
#     largest_column_count = 0
#     # target_row = 'RMS'
#     row_target = 0
#     # target_col = 'Mean\''
#     col_target = 0
#
#     with open(path, 'r') as f:
#         # reader = csv.reader(f)
#         lines = f.readlines()
#         # print(lines)
#         for line in lines:
#             column_count = len(line.split(',')) + 1
#             largest_column_count = column_count if largest_column_count < column_count else largest_column_count
#     f.close()
#
#     column_names = [i for i in range(0, largest_column_count)]
#
#     df = pd.read_csv(path, header=None, delimiter=',', names=column_names)
#     row_count = df.shape[0]
#     col_count = df.shape[1]
#     print(row_count, col_count)
#
#     for row in range(0, int(row_count)):
#         for col in range(0, int(col_count)):
#             if df[col][row] == target_row:
#                 row_target = row
#                 # print(row_target)
#             if df[col][row] == target_col:
#                 col_target = col
#     # print(row_target, col_target)
#     mea_data = df[col_target][row_target]
#     return mea_data
#


# handle csv files to get the measured data
def csv_handle_new(path, target_row, target_col):

    largest_column_count = 0
    row_target = 0
    col_target = 0

    with open(path, 'r') as f:
        # reader = csv.reader(f)
        lines = f.readlines()
        # print(lines)
        for line in lines:
            column_count = len(line.split(','))
            largest_column_count = column_count if largest_column_count < column_count else largest_column_count
    f.close()

    column_names = [i for i in range(0, largest_column_count)]

    df = pd.read_csv(path, header=None, delimiter=',', names=column_names)
    row_count = df.shape[0]
    col_count = df.shape[1]
    # print(row_count, col_count)

    # for row in range(0, int(row_count)):
    #     for col in range(0, int(col_count)):
    #
    #         if df[col][row] == target_row:
    #             row_target = row
    #             # print(row_target)
    #         if df[col][row] == target_col:
    #             col_target = col
    # # print(row_target, col_target)
    mea_data = df[target_col][target_row]
    return mea_data


def csv_shape(path):

    largest_column_count = 0
    row_target = 0
    col_target = 0

    with open(path, 'r') as f:
        # reader = csv.reader(f)
        lines = f.readlines()
        # print(lines)
        for line in lines:
            column_count = len(line.split(',')) + 1
            largest_column_count = column_count if largest_column_count < column_count else largest_column_count
    f.close()

    column_names = [i for i in range(0, largest_column_count)]

    df = pd.read_csv(path, header=None, delimiter=',', names=column_names)
    row_count = df.shape[0]
    col_count = df.shape[1]
    return row_count, col_count


# scan the test plan and get the row of test id
def get_id_row(target_id):

    id_row_number = 0
    wb = openpyxl.load_workbook(r'C:\Users\qwj\Desktop\Data Processing\test plan\2 [TPDR-05-1-001]R0A I2C test plan_v1.0.XLSX')
    # wb = openpyxl.load_workbook(file)
    sheet = wb['I2C']
    # sheet = wb['Sheet1']
    rows = sheet.max_row
    # print(rows)
    columns = sheet.max_column
    # print(columns)

    for row in range(1, rows+1):
        for column in range(1, columns+1):
            value = sheet.cell(row, column).value

            if str(value) == target_id:
                id_row_number = row
                # print(id_row_number)
    return id_row_number


rootPath = os.getcwd()
csvPath = os.path.join(rootPath, 'test data', 'I2C', 'csv')
# scan the id number in the csv folder
id_list = os.listdir(csvPath)
print(id_list)

# prepare the test plan for data collection
planPath = os.path.join(rootPath, 'test plan', '2 [TPDR-05-1-001]R0A I2C test plan_v1.0.xlsx')
wb = openpyxl.load_workbook(planPath)
sheet = wb['I2C']
m_row = sheet.max_row
m_col = sheet.max_column

fill_yellow = PatternFill('solid', fgColor='f8c600')  # Yellow color
fill_red = PatternFill('solid', fgColor='ff0000')  # Red color
font_blue = Font(color=colors.BLUE)  # blue color
font_red = Font(color='FF0000')  # red color
# font_yellow = Font(color='')
align = Alignment(horizontal='center', vertical='center')


# check the test items in every id folder
for idn in id_list:

    for row in sheet.iter_rows(min_row=1, max_row=m_row, min_col=1, max_col=m_col):
        for cell in row:
            sub_str = 'Sub-case ID:' + idn
            if sub_str == cell.value:
                print(cell.coordinate)
                id_row = cell.row

    idn_path = os.path.join(csvPath, idn)
    item_list = os.listdir(idn_path)
    num_csv = len(item_list)
    print(idn, num_csv, item_list)
    for item in item_list:
        item_Path = os.path.join(idn_path, item)
        print(item_Path)

    # get the voltage value and keep 3 bits after '.'
    if int(num_csv) == 7:
        if 'V.csv' in item_list:
            v_Path = os.path.join(idn_path, 'V.csv')
            SCL_Max = round(float(csv_handle_new(v_Path, 0, 1)), 3)
            SCL_Min = round(float(csv_handle_new(v_Path, 1, 1)), 3)
            SCL_H = round(float(csv_handle_new(v_Path, 2, 1)), 3)
            SCL_L = round(float(csv_handle_new(v_Path, 3, 1)), 3)
            SDA_Max = round(float(csv_handle_new(v_Path, 4, 1)), 3)
            SDA_Min = round(float(csv_handle_new(v_Path, 5, 1)), 3)
            SDA_H = round(float(csv_handle_new(v_Path, 6, 1)), 3)
            SDA_L = round(float(csv_handle_new(v_Path, 7, 1)), 3)
            print(SCL_Max, SCL_Min, SCL_H, SCL_L, SDA_Max, SDA_Min, SDA_H, SDA_L)

    if int(num_csv) == 6:
        if 'v.csv' in item_list:
            v_Path = os.path.join(idn_path, 'v.csv')
            SDA_Max = round(float(csv_handle_new(v_Path, 0, 1)), 3)
            SDA_Min = round(float(csv_handle_new(v_Path, 1, 1)), 3)
            SDA_H = round(float(csv_handle_new(v_Path, 2, 1)), 3)
            SDA_L = round(float(csv_handle_new(v_Path, 3, 1)), 3)
            print(SDA_Max, SDA_Min, SDA_H, SDA_L)

    # get frequency/high time/low time value and keep 3 bit after '.'.
    if 'Freq.csv' in item_list:
        f_Path = os.path.join(idn_path, 'Freq.csv')
        freq = round(float(csv_handle_new(f_Path, 0, 1)/1000), 3)
        high_T = round(float(csv_handle_new(f_Path, 1, 1)*1e6), 3)
        low_T = round(float(csv_handle_new(f_Path, 2, 1)*1e6), 3)
        print(freq, high_T, low_T)

    # get rise/fall time value and keep 3 bit after '.'.
    if 'RF.csv' in item_list:
        rf_Path = os.path.join(idn_path, 'RF.csv')
        row_c, col_c = csv_shape(rf_Path)
        if row_c > 2:
            SCL_R = round(float(csv_handle_new(rf_Path, 0, 1)*1e9), 3)
            SCL_F = round(float(csv_handle_new(rf_Path, 1, 1)*1e9), 3)
            SDA_R = round(float(csv_handle_new(rf_Path, 2, 1)*1e9), 3)
            SDA_F = round(float(csv_handle_new(rf_Path, 3, 1)*1e9), 3)
            print(SCL_R, SCL_F, SDA_R, SDA_F)
        else:
            SDA_R = round(float(csv_handle_new(rf_Path, 0, 1)*1e9), 3)
            SDA_F = round(float(csv_handle_new(rf_Path, 1, 1)*1e9), 3)
            print(SDA_R, SDA_F)

    # get setup time value and keep 3 bit after '.'.
    if 'Tsu.csv' in item_list:
        tsu_Path = os.path.join(idn_path, 'Tsu.csv')
        tsu_mea = round(float(csv_handle_new(tsu_Path, 0, 1)*1e9), 3)
        print(tsu_mea)

    # get hold time value and keep 3 bit after '.'.
    if 'Thd.csv' in item_list:
        thd_Path = os.path.join(idn_path, 'Thd.csv')
        thd_mea = round(float(csv_handle_new(thd_Path, 0, 1)*1e9), 3)
        print(thd_mea)

    # get start time value and keep 3 bit after '.'.
    if 'Sta.csv' in item_list:
        sta_Path = os.path.join(idn_path, 'Sta.csv')
        sta_mea = round(float(csv_handle_new(sta_Path, 0, 1)*1e6), 3)
        print(sta_mea)

    # get stop time value and keep 3 bit after '.'.
    if 'Sto.csv' in item_list:
        sto_Path = os.path.join(idn_path, 'Sto.csv')
        sto_mea = round(float(csv_handle_new(sto_Path, 0, 1)*1e6), 3)
        print(sto_mea)

    if int(num_csv) == 7:

        # input the voltage value to the test plan
        sheet.cell(id_row + 9, 5).value = SCL_Max
        sheet.cell(id_row + 10, 5).value = SCL_Min
        sheet.cell(id_row + 11, 5).value = SCL_H
        sheet.cell(id_row + 12, 5).value = SCL_L
        sheet.cell(id_row + 13, 5).value = SDA_Max
        sheet.cell(id_row + 14, 5).value = SDA_Min
        sheet.cell(id_row + 15, 5).value = SDA_H
        sheet.cell(id_row + 16, 5).value = SDA_L

    if int(num_csv) == 1:
        # input the voltage value to the test plan
        sheet.cell(id_row + 6, 5).value = SCL_Max
        sheet.cell(id_row + 7, 5).value = SCL_Min
        sheet.cell(id_row + 8, 5).value = SCL_H
        sheet.cell(id_row + 9, 5).value = SCL_L
        sheet.cell(id_row + 10, 5).value = SDA_Max
        sheet.cell(id_row + 11, 5).value = SDA_Min
        sheet.cell(id_row + 12, 5).value = SDA_H
        sheet.cell(id_row + 13, 5).value = SDA_L

    if int(num_csv) == 7:
        # input the timing value to the test plan
        sheet.cell(id_row + 6, 5).value = freq
        print(freq)
        sheet.cell(id_row + 7, 5).value = high_T
        sheet.cell(id_row + 8, 5).value = low_T

        sheet.cell(id_row + 17, 5).value = SCL_R
        sheet.cell(id_row + 18, 5).value = SCL_F
        sheet.cell(id_row + 19, 5).value = SDA_R
        sheet.cell(id_row + 20, 5).value = SDA_F

        sheet.cell(id_row + 21, 5).value = tsu_mea
        sheet.cell(id_row + 22, 5).value = thd_mea

        sheet.cell(id_row + 23, 5).value = sta_mea
        sheet.cell(id_row + 24, 5).value = sto_mea

        sheet.cell(id_row + 25, 5).value = 'Yes'

    if int(num_csv) == 6:

        sheet.cell(id_row + 6, 5).value = SDA_Max
        sheet.cell(id_row + 7, 5).value = SDA_Min
        sheet.cell(id_row + 8, 5).value = SDA_H
        sheet.cell(id_row + 9, 5).value = SDA_L

        sheet.cell(id_row + 10, 5).value = SDA_R
        sheet.cell(id_row + 11, 5).value = SDA_F

        sheet.cell(id_row + 12, 5).value = tsu_mea
        sheet.cell(id_row + 13, 5).value = thd_mea

        sheet.cell(id_row + 14, 5).value = sta_mea
        sheet.cell(id_row + 15, 5).value = sto_mea

    # margin calculation
    if int(num_csv) == 7:
        # SCL_MAX/Min/High/Low margin
        spec_min_SCL_Max, spec_max_SCL_Max = sheet.cell(id_row + 9, 3).value, sheet.cell(id_row + 9, 4).value
        if type(spec_max_SCL_Max) is int or float:
            sheet.cell(id_row + 9, 7).value = (spec_max_SCL_Max - sheet.cell(id_row + 9, 5).value) / spec_max_SCL_Max

        spec_min_SCL_Min, spec_max_SCL_Min = sheet.cell(id_row + 10, 3).value, sheet.cell(id_row + 10, 4).value
        if type(spec_min_SCL_Min) is int or float:
            sheet.cell(id_row + 10, 7).value = (spec_min_SCL_Min - sheet.cell(id_row + 10, 5).value) / spec_min_SCL_Min

        spec_min_SCL_H, spec_max_SCL_H = sheet.cell(id_row + 11, 3).value, sheet.cell(id_row + 11, 4).value
        if type(spec_min_SCL_H) is int or float:
            sheet.cell(id_row + 11, 7).value = (sheet.cell(id_row + 11, 5).value - spec_min_SCL_H) / spec_min_SCL_H

        spec_min_SCL_L, spec_max_SCL_L = sheet.cell(id_row + 12, 3).value, sheet.cell(id_row + 12, 4).value
        if type(spec_max_SCL_L) is int or float:
            sheet.cell(id_row + 12, 7).value = (spec_max_SCL_L - sheet.cell(id_row + 12, 5).value) / spec_max_SCL_L

        # SDA_MAX/Min/High/Low margin
        spec_min_SDA_Max, spec_max_SDA_Max = sheet.cell(id_row + 13, 3).value, sheet.cell(id_row + 13, 4).value
        if type(spec_max_SDA_Max) is int or float:
            sheet.cell(id_row + 13, 7).value = (spec_max_SDA_Max - sheet.cell(id_row + 13, 5).value) / spec_max_SDA_Max

        spec_min_SDA_Min, spec_max_SDA_Min = sheet.cell(id_row + 14, 3).value, sheet.cell(id_row + 14, 4).value
        if type(spec_min_SDA_Min) is int or float:
            sheet.cell(id_row + 14, 7).value = (spec_min_SDA_Min - sheet.cell(id_row + 14, 5).value) / spec_min_SDA_Min

        spec_min_SDA_H, spec_max_SDA_H = sheet.cell(id_row + 15, 3).value, sheet.cell(id_row + 15, 4).value
        if type(spec_min_SDA_H) is int or float:
            sheet.cell(id_row + 15, 7).value = (sheet.cell(id_row + 15, 5).value - spec_min_SDA_H) / spec_min_SDA_H

        spec_min_SDA_L, spec_max_SDA_L = sheet.cell(id_row + 16, 3).value, sheet.cell(id_row + 16, 4).value
        if type(spec_max_SDA_L) is int or float:
            sheet.cell(id_row + 16, 7).value = (spec_max_SDA_L - sheet.cell(id_row + 16, 5).value) / spec_max_SDA_L

        # frequency margin
        if sheet.cell(id_row + 6, 5).value < sheet.cell(id_row + 6, 4).value:
            sheet.cell(id_row + 6, 7).value = '/'
        #
        # # high time margin
        spec_min_H_time = sheet.cell(id_row + 7, 3).value
        spec_max_H_time = sheet.cell(id_row + 7, 4).value
        if type(spec_min_H_time) is int or float:
            sheet.cell(id_row + 7, 7).value = (sheet.cell(id_row + 7, 5).value - spec_min_H_time) / spec_min_H_time

        # # low time margin
        spec_min_L_time = sheet.cell(id_row + 8, 3).value
        spec_max_L_time = sheet.cell(id_row + 8, 4).value
        if type(spec_min_L_time) is int or float:
            sheet.cell(id_row + 8, 7).value = (sheet.cell(id_row + 8, 5).value - spec_min_L_time) / spec_min_L_time

        # SCL_Rise margin
        spec_min_SCL_R = sheet.cell(id_row + 17, 3).value
        spec_max_SCL_R = sheet.cell(id_row + 17, 4).value
        min_margin = sheet.cell(id_row + 17, 5).value - spec_min_SCL_R
        max_margin = spec_max_SCL_R - sheet.cell(id_row + 17, 5).value
        if min_margin <= max_margin:
            sheet.cell(id_row + 17, 7).value = min_margin / spec_min_SCL_R
        else:
            sheet.cell(id_row + 17, 7).value = max_margin / spec_max_SCL_R

        # SCL_Fall margin
        spec_min_SCL_F = sheet.cell(id_row + 17, 2).value
        spec_max_SCL_F = sheet.cell(id_row + 17, 3).value
        min_margin_F = sheet.cell(id_row + 17, 4).value - sheet.cell(id_row + 17, 2).value
        max_margin_F = sheet.cell(id_row + 17, 3).value - sheet.cell(id_row + 17, 4).value
        if min_margin_F < max_margin_F:
            sheet.cell(id_row + 17, 6).value = '/'
        else:
            sheet.cell(id_row + 17, 6).value = max_margin_F / spec_max_SCL_F

        # SDA_Rise margin
        spec_min_SDA_R = sheet.cell(id_row + 18, 2).value
        spec_max_SDA_R = sheet.cell(id_row + 18, 3).value
        min_margin_SDA_R = sheet.cell(id_row + 18, 4).value - sheet.cell(id_row + 18, 2).value
        max_margin_SDA_R = sheet.cell(id_row + 18, 3).value - sheet.cell(id_row + 18, 4).value
        if min_margin_SDA_R <= max_margin_SDA_R:
            sheet.cell(id_row + 18, 6).value = min_margin_SDA_R / spec_min_SDA_R
        else:
            sheet.cell(id_row + 18, 6).value = max_margin_SDA_R / spec_max_SDA_R

        # SDA_Fall margin
        spec_min_SDA_F = sheet.cell(id_row + 19, 2).value
        spec_max_SDA_F = sheet.cell(id_row + 19, 3).value
        min_margin_SDA_F = sheet.cell(id_row + 19, 4).value - sheet.cell(id_row + 19, 2).value
        max_margin_SDA_F = sheet.cell(id_row + 19, 3).value - sheet.cell(id_row + 19, 4).value
        if min_margin_SDA_F < max_margin_SDA_F:
            sheet.cell(id_row + 19, 6).value = '/'
        else:
            sheet.cell(id_row + 19, 6).value = max_margin_SDA_F / spec_max_SDA_F

        # Tsu margin
        spec_min_tsu = sheet.cell(id_row + 20, 2).value
        spec_max_tsu = sheet.cell(id_row + 20, 3).value
        if type(spec_min_tsu) is int or float:
            sheet.cell(id_row + 20, 6).value = (sheet.cell(id_row + 20, 4).value - spec_min_tsu) / spec_min_tsu

        #
        # Thd margin
        spec_min_thd = sheet.cell(id_row + 21, 2).value
        spec_max_thd = sheet.cell(id_row + 21, 3).value
        if spec_min_thd == 0:
            sheet.cell(id_row + 21, 6).value = '/'
        else:
            sheet.cell(id_row + 21, 6).value = (sheet.cell(id_row + 21, 4).value - spec_min_thd) / spec_min_thd

        # Sta margin
        spec_min_Sta = sheet.cell(id_row + 22, 2).value
        spec_max_Sta = sheet.cell(id_row + 22, 3).value

        if type(spec_min_Sta) is int or float:
            sheet.cell(id_row + 22, 6).value = (sheet.cell(id_row + 22, 4).value - spec_min_Sta) / spec_min_Sta

        # Sto margin
        spec_min_Sto = sheet.cell(id_row + 23, 2).value
        spec_max_Sto = sheet.cell(id_row + 23, 3).value

        if type(spec_min_Sto) is int or float:
            sheet.cell(id_row + 23, 6).value = (sheet.cell(id_row + 23, 4).value - spec_min_Sto) / spec_min_Sto

        # SCL monotonic margin
        sheet.cell(id_row + 24, 6).value = '/'

    # # format
    # for row in range(id_row + 5, id_row + 25):
    #     sheet.cell(row, 6).number_format = '0.00%'
    #     sheet.cell(row, 6).alignment = align
    #     if sheet.cell(row, 6).value == '/':
    #         sheet.cell(row, 6).font = font_blue
    #         sheet.cell(row, 4).font = font_blue
    #         sheet.cell(row, 7).value = 'Pass'
    #     elif type(sheet.cell(row, 6).value) is int or float:
    #         if sheet.cell(row, 6).value >= 0.05:
    #             sheet.cell(row, 6).font = font_blue
    #             sheet.cell(row, 4).font = font_blue
    #             sheet.cell(row, 7).value = 'Pass'
    #         if 0 < sheet.cell(row, 6).value < 0.05:
    #             sheet.cell(row, 6).fill = fill_yellow
    #             sheet.cell(row, 6).fill = fill_yellow
    #             sheet.cell(row, 7).value = 'Marginal Pass'
    #         if sheet.cell(row, 6).value <= 0:
    #             sheet.cell(row, 6).fill = fill_red
    #             sheet.cell(row, 6).fill = fill_red
    #             sheet.cell(row, 7).value = 'Fail'
    #     sheet.cell(row, 7).alignment = align
    #
    # # insert the capture
    # img = Image("E:\Grace\Data Processing\\test data\I2C\waveform\\13.62\\freq.png")
    # img.width = 270
    # img.height = 170
    # sheet.add_image(img, 'A29')
    #
    # img = Image("E:\Grace\Data Processing\\test data\I2C\waveform\\13.62\\v.png")
    # img.width = 270
    # img.height = 170
    # sheet.add_image(img, 'D29')
    #
    # img = Image("E:\Grace\Data Processing\\test data\I2C\waveform\\13.62\\rf.png")
    # img.width = 270
    # img.height = 170
    # sheet.add_image(img, 'A31')
    #
    # img = Image("E:\Grace\Data Processing\\test data\I2C\waveform\\13.62\\tsu.png")
    # img.width = 270
    # img.height = 170
    # sheet.add_image(img, 'D31')
    #
    # img = Image("E:\Grace\Data Processing\\test data\I2C\waveform\\13.62\\thd.png")
    # img.width = 270
    # img.height = 170
    # sheet.add_image(img, 'A33')
    #
    # img = Image("E:\Grace\Data Processing\\test data\I2C\waveform\\13.62\\sta.png")
    # img.width = 270
    # img.height = 170
    # sheet.add_image(img, 'D33')
    #
    # img = Image("E:\Grace\Data Processing\\test data\I2C\waveform\\13.62\\sto.png")
    # img.width = 270
    # img.height = 170
    # sheet.add_image(img, 'A35')
    #
    # img = Image("E:\Grace\Data Processing\\test data\I2C\waveform\\13.62\\r.png")
    # img.width = 270
    # img.height = 170
    # sheet.add_image(img, 'D35')
    #
    # img = Image("E:\Grace\Data Processing\\test data\I2C\waveform\\13.62\\f.png")
    # img.width = 270
    # img.height = 170
    # sheet.add_image(img, 'A37')


    # # margin calculation
    # if int(num_csv) == 1:
    #     # SCL_MAX&SDA_MAX margin
    #     if type(sheet.cell(id_row, 10).value) is int or float:
    #         sheet.cell(id_row, 12).value = (sheet.cell(id_row, 10).value - sheet.cell(id_row, 11).value) / sheet.cell(
    #         id_row, 10).value
    #         sheet.cell(id_row + 1, 12).value = (sheet.cell(id_row, 10).value - sheet.cell(id_row + 1, 11).value) / \
    #                                            sheet.cell(id_row, 10).value
    #         sheet.cell(id_row, 12).number_format = '0.00%'
    #         sheet.cell(id_row + 1, 12).number_format = '0.00%'

    #
    #     elif sheet.cell(id_row, 10).value == 'NA':
    #         sheet.cell(id_row, 12).value = '/'
    #         sheet.cell(id_row + 1, 12).value = '/'
    #         sheet.cell(id_row, 12).font = font_blue
    #         sheet.cell(id_row + 1, 12).font = font_blue
    #
    #     # SCL_Min&SDA_Min margin
    #     if type(sheet.cell(id_row, 13).value) is int or float:
    #         sheet.cell(id_row, 15).value = (sheet.cell(id_row, 13).value - sheet.cell(id_row, 14).value) / sheet.cell(
    #         id_row, 13).value
    #         sheet.cell(id_row + 1, 15).value = (sheet.cell(id_row, 13).value - sheet.cell(id_row + 1, 14).value) / \
    #                                            sheet.cell(id_row, 13).value
    #         sheet.cell(id_row, 15).number_format = '0.00%'
    #         sheet.cell(id_row + 1, 15).number_format = '0.00%'
    #         # set the cell color
    #         if sheet.cell(id_row, 15).value >= 0.05:
    #             sheet.cell(id_row, 15).font = font_blue
    #             sheet.cell(id_row, 14).font = font_blue
    #         if 0 < sheet.cell(id_row, 15).value < 0.05:
    #             sheet.cell(id_row, 15).fill = fill_yellow
    #             sheet.cell(id_row, 14).fill = fill_yellow
    #         if sheet.cell(id_row, 15).value <= 0:
    #             sheet.cell(id_row, 15).fill = fill_red
    #             sheet.cell(id_row, 14).fill = fill_red
    #
    #         if sheet.cell(id_row + 1, 15).value >= 0.05:
    #             sheet.cell(id_row + 1, 15).font = font_blue
    #             sheet.cell(id_row + 1, 14).font = font_blue
    #         if 0 < sheet.cell(id_row + 1, 15).value < 0.05:
    #             sheet.cell(id_row + 1, 15).fill = fill_yellow
    #             sheet.cell(id_row + 1, 14).fill = fill_yellow
    #         if sheet.cell(id_row + 1, 15).value <= 0:
    #             sheet.cell(id_row + 1, 15).fill = fill_red
    #             sheet.cell(id_row + 1, 14).fill = fill_red
    #
    #     elif sheet.cell(id_row, 13).value == 'NA':
    #         sheet.cell(id_row, 15).value = '/'
    #         sheet.cell(id_row + 1, 15).value = '/'
    #         sheet.cell(id_row, 15).font = font_blue
    #         sheet.cell(id_row + 1, 15).font = font_blue
    #
    #     # SCL_H margin
    #     # if type(sheet.cell(id_row, 16).value) is int or float:
    #     #     sheet.cell(id_row, 19).value = (sheet.cell(id_row, 18).value - sheet.cell(id_row, 16).value) / sheet.cell(
    #     #         id_row, 16).value
    #     # elif sheet.cell(id_row, 16).value == 'NA':
    #     #     sheet.cell(id_row, 19).value = '/'
    #     sheet.cell(id_row, 19).value = '/'
    #     sheet.cell(id_row, 19).font = font_blue
    #     sheet.cell(id_row, 18).font = font_blue
    #
    #     # SDA_H margin
    #     if type(sheet.cell(id_row + 1, 16).value) is int or float:
    #         sheet.cell(id_row + 1, 19).value = (sheet.cell(id_row + 1, 18).value - sheet.cell(id_row + 1, 16).value) / \
    #                                            sheet.cell(id_row + 1, 16).value
    #         sheet.cell(id_row + 1, 19).number_format = '0.00%'
    #         # set the cell color
    #         if sheet.cell(id_row + 1, 19).value >= 0.05:
    #             sheet.cell(id_row + 1, 19).font = font_blue
    #             sheet.cell(id_row + 1, 18).font = font_blue
    #         if 0 < sheet.cell(id_row + 1, 19).value < 0.05:
    #             sheet.cell(id_row + 1, 19).fill = fill_yellow
    #             sheet.cell(id_row + 1, 18).fill = fill_yellow
    #         if sheet.cell(id_row + 1, 19).value <= 0:
    #             sheet.cell(id_row + 1, 19).fill = fill_red
    #             sheet.cell(id_row + 1, 18).fill = fill_red
    #
    #     elif sheet.cell(id_row + 1, 16).value == 'NA':
    #         sheet.cell(id_row + 1, 19).value = '/'
    #         sheet.cell(id_row + 1, 19).font = font_blue
    #         sheet.cell(id_row + 1, 18).font = font_blue
    #
    #     # SCL_L margin
    #     if type(sheet.cell(id_row, 21).value) is int or float:
    #         sheet.cell(id_row, 23).value = (sheet.cell(id_row, 21).value - sheet.cell(id_row, 22).value) / sheet.cell(
    #         id_row, 21).value
    #         sheet.cell(id_row, 23).number_format = '0.00%'
    #         # set the cell color
    #         if sheet.cell(id_row, 23).value >= 0.05:
    #             sheet.cell(id_row, 23).font = font_blue
    #             sheet.cell(id_row, 22).font = font_blue
    #         if 0 < sheet.cell(id_row, 23).value < 0.05:
    #             sheet.cell(id_row, 23).fill = fill_yellow
    #             sheet.cell(id_row, 22).fill = fill_yellow
    #         if sheet.cell(id_row, 23).value <= 0:
    #             sheet.cell(id_row, 23).fill = fill_red
    #             sheet.cell(id_row, 22).fill = fill_red
    #
    #     elif sheet.cell(id_row, 21).value == 'NA':
    #         sheet.cell(id_row, 23).value = '/'
    #         sheet.cell(id_row, 23).font = font_blue
    #
    #     # SDA_L margin
    #     if type(sheet.cell(id_row + 1, 21).value) is int or float:
    #         sheet.cell(id_row + 1, 23).value = (sheet.cell(id_row + 1, 21).value - sheet.cell(id_row + 1, 22).value) / \
    #                                            sheet.cell(id_row + 1, 21).value
    #         sheet.cell(id_row + 1, 23).number_format = '0.00%'
    #         # set the cell color
    #         if sheet.cell(id_row + 1, 23).value >= 0.05:
    #             sheet.cell(id_row + 1, 23).font = font_blue
    #             sheet.cell(id_row + 1, 22).font = font_blue
    #         if 0 < sheet.cell(id_row + 1, 23).value < 0.05:
    #             sheet.cell(id_row + 1, 23).fill = fill_yellow
    #             sheet.cell(id_row + 1, 22).fill = fill_yellow
    #         if sheet.cell(id_row + 1, 23).value <= 0:
    #             sheet.cell(id_row + 1, 23).fill = fill_red
    #             sheet.cell(id_row + 1, 22).fill = fill_red
    #
    #     elif sheet.cell(id_row + 1, 21).value == 'NA':
    #         sheet.cell(id_row + 1, 23).value = '/'
    #         sheet.cell(id_row + 1, 23).font = font_blue
    #
    # # margin calculation
    # if int(num_csv) == 6:
    #     # SDA_Max margin
    #     if type(sheet.cell(id_row + 1, 10).value) is int or float:
    #         sheet.cell(id_row + 1, 12).value = (sheet.cell(id_row + 1, 10).value - sheet.cell(id_row + 1, 11).value)/\
    #                                            sheet.cell(id_row + 1, 10).value
    #         sheet.cell(id_row + 1, 12).number_format = '0.00%'
    #         # set the cell color
    #         if sheet.cell(id_row + 1, 12).value >= 0.05:
    #             sheet.cell(id_row + 1, 12).font = font_blue
    #             sheet.cell(id_row + 1, 11).font = font_blue
    #         if 0 < sheet.cell(id_row + 1, 12).value < 0.05:
    #             sheet.cell(id_row + 1, 12).fill = fill_yellow
    #             sheet.cell(id_row + 1, 11).fill = fill_yellow
    #         if sheet.cell(id_row + 1, 12).value <= 0:
    #             sheet.cell(id_row + 1, 12).fill = fill_red
    #             sheet.cell(id_row + 1, 11).fill = fill_red
    #
    #     elif sheet.cell(id_row + 1, 10).value == 'NA':
    #         sheet.cell(id_row + 1, 12).value = '/'
    #         sheet.cell(id_row + 1, 12).font = font_blue
    #         sheet.cell(id_row + 1, 11).font = font_blue
    #
    #     # SDA_Min margin
    #     if type(sheet.cell(id_row + 1, 13).value) is int or float:
    #         sheet.cell(id_row + 1, 15).value = (sheet.cell(id_row + 1, 13).value - sheet.cell(id_row + 1, 14).value) / \
    #                                            sheet.cell(id_row + 1, 13).value
    #         sheet.cell(id_row + 1, 15).number_format = '0.00%'
    #         # set the cell color
    #         if sheet.cell(id_row + 1, 15).value >= 0.05:
    #             sheet.cell(id_row + 1, 15).font = font_blue
    #             sheet.cell(id_row + 1, 14).font = font_blue
    #         if 0 < sheet.cell(id_row + 1, 15).value < 0.05:
    #             sheet.cell(id_row + 1, 15).fill = fill_yellow
    #             sheet.cell(id_row + 1, 14).fill = fill_yellow
    #         if sheet.cell(id_row + 1, 15).value <= 0:
    #             sheet.cell(id_row + 1, 15).fill = fill_red
    #             sheet.cell(id_row + 1, 14).fill = fill_red
    #
    #     elif sheet.cell(id_row + 1, 13).value == 'NA':
    #         sheet.cell(id_row + 1, 15).value = '/'
    #         sheet.cell(id_row + 1, 15).font = font_blue
    #         sheet.cell(id_row + 1, 14).font = font_blue
    #
    #     # SDA_H margin
    #     if type(sheet.cell(id_row + 1, 16).value) is int or float:
    #         sheet.cell(id_row + 1, 19).value = (sheet.cell(id_row + 1, 18).value - sheet.cell(id_row + 1, 16).value) / \
    #                                            sheet.cell(id_row + 1, 16).value
    #         sheet.cell(id_row + 1, 19).number_format = '0.00%'
    #         # set the cell color
    #         if sheet.cell(id_row + 1, 19).value >= 0.05:
    #             sheet.cell(id_row + 1, 19).font = font_blue
    #             sheet.cell(id_row + 1, 18).font = font_blue
    #         if 0 < sheet.cell(id_row + 1, 19).value < 0.05:
    #             sheet.cell(id_row + 1, 19).fill = fill_yellow
    #             sheet.cell(id_row + 1, 18).fill = fill_yellow
    #         if sheet.cell(id_row + 1, 19).value <= 0:
    #             sheet.cell(id_row + 1, 19).fill = fill_red
    #             sheet.cell(id_row + 1, 18).fill = fill_red
    #
    #     elif sheet.cell(id_row + 1, 16).value == 'NA':
    #         sheet.cell(id_row + 1, 19).value = '/'
    #         sheet.cell(id_row + 1, 19).font = font_blue
    #         sheet.cell(id_row + 1, 18).font = font_blue
    #
    #     # SDA_L margin
    #     if type(sheet.cell(id_row + 1, 21).value) is int or float:
    #         sheet.cell(id_row + 1, 23).value = (sheet.cell(id_row + 1, 21).value - sheet.cell(id_row + 1, 22).value) / \
    #                                            sheet.cell(id_row + 1, 21).value
    #         sheet.cell(id_row + 1, 23).number_format = '0.00%'
    #         # set the cell color
    #         if sheet.cell(id_row + 1, 23).value >= 0.05:
    #             sheet.cell(id_row + 1, 23).font = font_blue
    #             sheet.cell(id_row + 1, 22).font = font_blue
    #         if 0 < sheet.cell(id_row + 1, 23).value < 0.05:
    #             sheet.cell(id_row + 1, 23).fill = fill_yellow
    #             sheet.cell(id_row + 1, 22).fill = fill_yellow
    #         if sheet.cell(id_row + 1, 23).value <= 0:
    #             sheet.cell(id_row + 1, 23).fill = fill_red
    #             sheet.cell(id_row + 1, 22).fill = fill_red
    #
    #     elif sheet.cell(id_row + 1, 21).value == 'NA':
    #         sheet.cell(id_row + 1, 23).value = '/'
    #         sheet.cell(id_row + 1, 23).font = font_blue
    #         sheet.cell(id_row + 1, 22).font = font_blue
    #
    #     # SDA_R margin
    #     if type(sheet.cell(id_row, 42).value) is int or float:  # SDA_Rise_max spec is a data
    #         if type(sheet.cell(id_row, 41).value) is int or float:  # SDA_Rise_min spec is a data
    #             if sheet.cell(id_row, 41).value < sheet.cell(id_row, 43).value < sheet.cell(id_row, 42).value:
    #                 # calculate the margin with the min spec
    #                 sheet.cell(id_row, 44).value = (sheet.cell(id_row, 42).value - sheet.cell(id_row, 43).value) / \
    #                                        sheet.cell(id_row, 42).value
    #                 sheet.cell(id_row, 44).number_format = '0.00%'
    #             # set the cell color
    #                 if sheet.cell(id_row, 44).value >= 0.05:
    #                     sheet.cell(id_row, 44).font = font_blue
    #                     sheet.cell(id_row, 43).font = font_blue
    #                 if 0 < sheet.cell(id_row, 44).value < 0.05:
    #                     sheet.cell(id_row, 44).fill = fill_yellow
    #                     sheet.cell(id_row, 43).fill = fill_yellow
    #             else:
    #                 sheet.cell(id_row, 44).value = '/'
    #                 sheet.cell(id_row, 44).fill = fill_red
    #                 sheet.cell(id_row, 43).fill = fill_red
    #
    #     # SDA_F margin
    #     if type(sheet.cell(id_row, 46).value) is int or float:
    #         sheet.cell(id_row, 48).value = (sheet.cell(id_row, 46).value - sheet.cell(id_row, 47).value) / \
    #                                        sheet.cell(id_row, 46).value
    #         sheet.cell(id_row, 48).number_format = '0.00%'
    #         # set the cell color
    #         if sheet.cell(id_row, 48).value >= 0.05:
    #             sheet.cell(id_row, 48).font = font_blue
    #             sheet.cell(id_row, 47).font = font_blue
    #         if 0 < sheet.cell(id_row, 48).value < 0.05:
    #             sheet.cell(id_row, 48).fill = fill_yellow
    #             sheet.cell(id_row, 47).fill = fill_yellow
    #         if sheet.cell(id_row, 48).value <= 0:
    #             sheet.cell(id_row, 48).fill = fill_red
    #             sheet.cell(id_row, 47).fill = fill_red
    #
    #     # Tsu margin
    #     if type(sheet.cell(id_row, 49).value) is int or float:
    #         sheet.cell(id_row, 51).value = (sheet.cell(id_row, 50).value - sheet.cell(id_row, 49).value) / \
    #                                        sheet.cell(id_row, 49).value
    #         sheet.cell(id_row, 51).number_format = '0.00%'
    #         # set the cell color
    #         if sheet.cell(id_row, 51).value >= 0.05:
    #             sheet.cell(id_row, 51).font = font_blue
    #             sheet.cell(id_row, 50).font = font_blue
    #         if 0 < sheet.cell(id_row, 51).value < 0.05:
    #             sheet.cell(id_row, 51).fill = fill_yellow
    #             sheet.cell(id_row, 50).fill = fill_yellow
    #         if sheet.cell(id_row, 51).value <= 0:
    #             sheet.cell(id_row, 51).fill = fill_red
    #             sheet.cell(id_row, 50).fill = fill_red
    #
    #     # Thd margin
    #     if sheet.cell(id_row, 52).value == 0:
    #         sheet.cell(id_row, 54).value = '/'
    #         sheet.cell(id_row, 54).font = font_blue
    #         sheet.cell(id_row, 53).font = font_blue
    #
    #     # Sta margin
    #     if type(sheet.cell(id_row, 55).value) is int or float:
    #         sheet.cell(id_row, 57).value = (sheet.cell(id_row, 56).value - sheet.cell(id_row, 55).value) / \
    #                                        sheet.cell(id_row, 55).value
    #         sheet.cell(id_row, 57).number_format = '0.00%'
    #         # set the cell color
    #         if sheet.cell(id_row, 57).value >= 0.05:
    #             sheet.cell(id_row, 57).font = font_blue
    #             sheet.cell(id_row, 56).font = font_blue
    #         if 0 < sheet.cell(id_row, 57).value < 0.05:
    #             sheet.cell(id_row, 57).fill = fill_yellow
    #             sheet.cell(id_row, 56).fill = fill_yellow
    #         if sheet.cell(id_row, 57).value <= 0:
    #             sheet.cell(id_row, 57).fill = fill_red
    #             sheet.cell(id_row, 56).fill = fill_red
    #
    #     # Sto margin
    #     if type(sheet.cell(id_row, 58).value) is int or float:
    #         sheet.cell(id_row, 60).value = (sheet.cell(id_row, 59).value - sheet.cell(id_row, 58).value) / \
    #                                        sheet.cell(id_row, 58).value
    #         sheet.cell(id_row, 60).number_format = '0.00%'
    #         # set the cell color
    #         if sheet.cell(id_row, 60).value >= 0.05:
    #             sheet.cell(id_row, 60).font = font_blue
    #             sheet.cell(id_row, 59).font = font_blue
    #         if 0 < sheet.cell(id_row, 60).value < 0.05:
    #             sheet.cell(id_row, 60).fill = fill_yellow
    #             sheet.cell(id_row, 59).fill = fill_yellow
    #         if sheet.cell(id_row, 60).value <= 0:
    #             sheet.cell(id_row, 60).fill = fill_red
    #             sheet.cell(id_row, 59).fill = fill_red

    # format the report

wb.save('I2C report.xlsx')
#


