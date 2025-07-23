# -*-coding:utf-8 -*-
# Author: Grace Qian
# Date: 2024-2-28

import os
import openpyxl
import pandas as pd
import csv
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import colors

# handle csv files to get the measured data
def csv_handle(path, target_row, target_col):

    largest_column_count = 0
    # target_row = 'RMS'
    row_target = 0
    # target_col = 'Mean\''
    col_target = 0

    with open(path, 'r') as f:
        # reader = csv.reader(f)
        lines = f.readlines()
        # print(lines)
        for line in lines:
            column_count = len(line.split(',')) + 1
            largest_column_count = column_count if largest_column_count < column_count else largest_column_count
    f.close()

    column_names = [i for i in range(0, largest_column_count)]

    df = pd.read_csv(path, header=None, delimiter=',', names=column_names)
    row_count = df.shape[0]
    col_count = df.shape[1]
    print(row_count, col_count)

    for row in range(0, int(row_count)):
        for col in range(0, int(col_count)):
            if df[col][row] == target_row:
                row_target = row
                # print(row_target)
            if df[col][row] == target_col:
                col_target = col
    # print(row_target, col_target)
    mea_data = df[col_target][row_target]
    return mea_data


# handle csv files to get the measured data
def csv_handle_new(path, target_row, target_col):

    largest_column_count = 0

    with open(path, 'r') as f:
        # reader = csv.reader(f)
        lines = f.readlines()
        # print(lines)
        for line in lines:
            column_count = len(line.split(',')) + 1
            largest_column_count = column_count if largest_column_count < column_count else largest_column_count
    f.close()
    # print(largest_column_count)
    column_names = [i for i in range(0, largest_column_count)]

    # df = pd.read_csv(path, header=None, delimiter=',', names=column_names)
    df = pd.read_csv(path, header=None, delimiter=',')

    # row_count = df.shape[0]
    # col_count = df.shape[1]
    # print(row_count, col_count)

    # for row in range(0, int(row_count)):
    #     for col in range(0, int(col_count)):
    #
    #         if df[col][row] == target_row:
    #             row_target = row
    #             # print(row_target)
    #         if df[col][row] == target_col:
    #             col_target = col
    # # print(row_target, col_target)
    mea_data = df[target_col][target_row]
    return mea_data


# def csv_shape(path):
#
#     largest_column_count = 0
#
#     with open(path, 'r') as f:
#         # reader = csv.reader(f)
#         lines = f.readlines()
#         # print(lines)
#         for line in lines:
#             column_count = len(line.split(',')) + 1
#             largest_column_count = column_count if largest_column_count < column_count else largest_column_count
#     f.close()
#
#     column_names = [i for i in range(0, largest_column_count)]
#
#     df = pd.read_csv(path, header=None, delimiter=',', names=column_names)
#     row_count = df.shape[0]
#     col_count = df.shape[1]
#     return row_count, col_count


# scan the test plan and get the row of test id
def get_id_row(wsheet, target_id):

    id_row_number = 0
    # wb = openpyxl.load_workbook(r'E:\Grace\Data Processing\test plan\13 I2C.xlsx')
    # wb = openpyxl.load_workbook(file)
    # sheet = wb['I2C']
    # sheet = wb['Sheet1']
    rows = wsheet.max_row
    # print(rows)
    columns = sheet.max_column
    # print(columns)

    for row in range(1, rows+1):
        for column in range(1, columns+1):
            value = sheet.cell(row, column).value

            if str(value) == target_id:
                id_row_number = row
                # print(id_row_number)
    return id_row_number


# get SCL&SDA voltage value and keep 3 bits after '.'
# def get_SCL_SDA_voltage(path):
#
#     SCL_Max_v = round(float(csv_handle_new(path, 5, 5)), 3)
#     SCL_Min_v = round(float(csv_handle_new(path, 7, 5)), 3)
#     SCL_H_v = round(float(csv_handle_new(path, 9, 5)), 3)
#     SCL_L_v = round(float(csv_handle_new(path, 11, 5)), 3)
#     SDA_Max_v = round(float(csv_handle_new(path, 13, 5)), 3)
#     SDA_Min_v = round(float(csv_handle_new(path, 15, 5)), 3)
#     SDA_H_v = round(float(csv_handle_new(path, 17, 5)), 3)
#     SDA_L_v = round(float(csv_handle_new(path, 19, 5)), 3)
#     return SCL_Max_v, SCL_Min_v, SCL_H_v, SCL_L_v, SDA_Max_v, SDA_Min_v, SDA_H_v, SDA_L_v

def get_SCL_SDA_voltage_6Series(path):

    SCL_Max_v = round(float(csv_handle_new(path, 5, 5)), 3)
    SCL_Min_v = round(float(csv_handle_new(path, 7, 5)), 3)
    SCL_H_v = round(float(csv_handle_new(path, 9, 5)), 3)
    SCL_L_v = round(float(csv_handle_new(path, 11, 5)), 3)
    SDA_Max_v = round(float(csv_handle_new(path, 13, 5)), 3)
    SDA_Min_v = round(float(csv_handle_new(path, 15, 5)), 3)
    SDA_H_v = round(float(csv_handle_new(path, 17, 5)), 3)
    SDA_L_v = round(float(csv_handle_new(path, 19, 5)), 3)
    return SCL_Max_v, SCL_Min_v, SCL_H_v, SCL_L_v, SDA_Max_v, SDA_Min_v, SDA_H_v, SDA_L_v

# get SDA voltage value and keep 3 bits after '.'
# def get_SDA_voltage(path):
#
#     SDA_Max_v = round(float(csv_handle_new(path, 0, 1)), 3)
#     SDA_Min_v = round(float(csv_handle_new(path, 1, 1)), 3)
#     SDA_H_v = round(float(csv_handle_new(path, 2, 1)), 3)
#     SDA_L_v = round(float(csv_handle_new(path, 3, 1)), 3)
#     return SDA_Max_v, SDA_Min_v, SDA_H_v, SDA_L_v


def get_SDA_voltage_6Series(path):

    SDA_Max_v = round(float((csv_handle_new(path, 5, 5)).split(' ')[0]), 3)
    SDA_Min_v = round(float((csv_handle_new(path, 7, 5)).split(' ')[0]), 3)
    SDA_H_v = round(float((csv_handle_new(path, 9, 5)).split(' ')[0]), 3)
    SDA_L_v = round(float((csv_handle_new(path, 11, 5)).split(' ')[0]), 3)
    return SDA_Max_v, SDA_Min_v, SDA_H_v, SDA_L_v

def get_freq(path):

    freq_t = round(float(csv_handle_new(path, 0, 1) / 1000), 3)
    high_T_t = round(float(csv_handle_new(path, 1, 1) * 1e6), 3)
    low_T_t = round(float(csv_handle_new(path, 2, 1) * 1e6), 3)
    return freq_t, high_T_t, low_T_t


# get the rise/fall time of SCL and SDA
# def get_SCL_SDA_RF(path):
#
#     SCL_R_t = round(float(csv_handle_new(path, 0, 1) * 1e9), 3)
#     SCL_F_t = round(float(csv_handle_new(path, 1, 1) * 1e9), 3)
#     SDA_R_t = round(float(csv_handle_new(path, 2, 1) * 1e9), 3)
#     SDA_F_t = round(float(csv_handle_new(path, 3, 1) * 1e9), 3)
#     return SCL_R_t, SCL_F_t, SDA_R_t, SDA_F_t


def get_SCL_SDA_RF_6Series(path):

    SCL_R_t = round(float(csv_handle_new(path, 0, 1) * 1e9), 3)
    SCL_F_t = round(float(csv_handle_new(path, 1, 1) * 1e9), 3)
    SDA_R_t = round(float(csv_handle_new(path, 2, 1) * 1e9), 3)
    SDA_F_t = round(float(csv_handle_new(path, 3, 1) * 1e9), 3)
    return SCL_R_t, SCL_F_t, SDA_R_t, SDA_F_t


# get the rise/fall time of SDA
# def get_SDA_RF(path):
#
#     SDA_R_t = round(float(csv_handle_new(path, 0, 1) * 1e9), 3)
#     SDA_F_t = round(float(csv_handle_new(path, 1, 1) * 1e9), 3)
#     return SDA_R_t, SDA_F_t


def get_SDA_RF_6Series(path):

    SDA_R_t = round(float(csv_handle_new(path, 5, 5).split(' ')), 3)
    SDA_F_t = round(float(csv_handle_new(path, 6, 5).split(' ')), 3)
    return SDA_R_t, SDA_F_t

# get setup time and hold time
# def get_tsu_thd(path):
#
#     tsu_thd_t = round(float(csv_handle_new(path, 0, 1)*1e9), 3)
#     return tsu_thd_t


def get_tsu_thd_6Series(path):

    tsu_thd_t = round(float(csv_handle_new(path, 5, 5).split(' ')[0]), 3)
    return tsu_thd_t


# get start and stop time
# def get_sta_sto(path):
#
#     sta_sto_t = round(float(csv_handle_new(path, 0, 1)*1e6), 3)
#     return sta_sto_t


def get_sta_sto_6Series(path):

    sta_sto_t = round(float(csv_handle_new(path, 5, 5).split(' ')[0]), 3)
    return sta_sto_t
#
# # input the voltage of SCL and SDA to test plan
# def set_SCL_SDA_voltage(wsheet, id, SCL_V_Max, SCL_V_Min, SCL_V_H, SCL_V_L, SDA_V_Max, SDA_V_Min, SDA_V_H, SDA_V_L):
#
#     wsheet.cell(id, 11).value = SCL_V_Max
#     wsheet.cell(id, 14).value = SCL_V_Min
#     wsheet.cell(id, 18).value = SCL_V_H
#     wsheet.cell(id, 22).value = SCL_V_L
#     wsheet.cell(id + 1, 11).value = SDA_V_Max
#     wsheet.cell(id + 1, 14).value = SDA_V_Min
#     wsheet.cell(id + 1, 18).value = SDA_V_H
#     wsheet.cell(id + 1, 22).value = SDA_V_L
#
#
# # input the voltage of SDA to the test plan
# def set_SDA_voltage(wsheet, id, SDA_V_Max, SDA_V_Min, SDA_V_H, SDA_V_L):
#
#     wsheet.cell(id + 1, 11).value = SDA_V_Max
#     wsheet.cell(id + 1, 14).value = SDA_V_Min
#     wsheet.cell(id + 1, 18).value = SDA_V_H
#     wsheet.cell(id + 1, 22).value = SDA_V_L


rootPath = os.getcwd()
csvPath = os.path.join(rootPath, 'test data', 'I2C', 'csv')

# scan the id number in the csv folder
id_list = os.listdir(csvPath)
print(id_list)

# prepare the test plan for data collection
planPath = os.path.join(rootPath, 'test plan', '13 I2C test plan_v0.3.xlsx')
wb = openpyxl.load_workbook(planPath)
sheet = wb['I2C']

fill_yellow = PatternFill('solid', fgColor='f8c600')  # Yellow color
fill_red = PatternFill('solid', fgColor='ff0000')  # Red color
font_blue = Font(color=colors.BLUE)  # blue color
font_red = Font(color='FF0000')  # red color
# font_yellow = Font(color='')


# check the test items in every id folder
for idn in id_list:

    idn_path = os.path.join(csvPath, idn)
    # check the test items in the id
    item_list = os.listdir(idn_path)
    # check the number of the test items in the id
    num_csv = len(item_list)
    print(idn, num_csv, item_list)
    print(idn + ' is processing the date...')

    # get the test data from csv files

    if int(num_csv) == 7:

        v_Path = os.path.join(idn_path, 'V.csv')
        SCL_Max, SCL_Min, SCL_H, SCL_L, SDA_Max, SDA_Min, SDA_H, SDA_L = get_SCL_SDA_voltage(v_Path)
        f_Path = os.path.join(idn_path, 'T.csv')
        freq, high_T, low_T = get_freq(f_Path)
        rf_Path = os.path.join(idn_path, 'RF.csv')
        SCL_R, SCL_F, SDA_R, SDA_F = get_SCL_SDA_RF(rf_Path)
        tsu_Path = os.path.join(idn_path, 'TSU.csv')
        tsu_mea = get_tsu_thd(tsu_Path)
        thd_Path = os.path.join(idn_path, 'THD.csv')
        thd_mea = get_tsu_thd(thd_Path)
        sta_Path = os.path.join(idn_path, 'STA.csv')
        sta_mea = get_sta_sto(sta_Path)
        sto_Path = os.path.join(idn_path, 'STO.csv')
        sto_mea = get_sta_sto(sto_Path)

    if int(num_csv) == 1:

        v_Path = os.path.join(idn_path, 'v.csv')
        SCL_Max, SCL_Min, SCL_H, SCL_L, SDA_Max, SDA_Min, SDA_H, SDA_L = get_SCL_SDA_voltage(v_Path)

    if int(num_csv) == 6:

        v_Path = os.path.join(idn_path, 'V.csv')
        SDA_Max, SDA_Min, SDA_H, SDA_L = get_SDA_voltage_6Series(v_Path)
        rf_Path = os.path.join(idn_path, 'RF.csv')
        SDA_R, SDA_F = get_SDA_RF_6Series(rf_Path)
        tsu_Path = os.path.join(idn_path, 'TSU.csv')
        tsu_mea = get_tsu_thd_6Series(tsu_Path)
        thd_Path = os.path.join(idn_path, 'THD.csv')
        thd_mea = get_tsu_thd_6Series(thd_Path)
        sta_Path = os.path.join(idn_path, 'STA.csv')
        sta_mea = get_sta_sto_6Series(sta_Path)
        sto_Path = os.path.join(idn_path, 'STO.csv')
        sto_mea = get_sta_sto_6Series(sto_Path)

    # input the test data to test plan

    id_row = get_id_row(sheet, idn)
    # print(id_row)
    if int(num_csv) == 7:

        # input the voltage value to the test plan
        sheet.cell(id_row, 11).value = SCL_Max
        sheet.cell(id_row, 14).value = SCL_Min
        sheet.cell(id_row, 18).value = SCL_H
        sheet.cell(id_row, 22).value = SCL_L

        sheet.cell(id_row + 1, 11).value = SDA_Max
        sheet.cell(id_row + 1, 14).value = SDA_Min
        sheet.cell(id_row + 1, 18).value = SDA_H
        sheet.cell(id_row + 1, 22).value = SDA_L

        # input the timing value to the test plan
        sheet.cell(id_row, 25).value = freq
        sheet.cell(id_row, 28).value = high_T
        sheet.cell(id_row, 31).value = low_T

        sheet.cell(id_row, 35).value = SCL_R
        sheet.cell(id_row, 39).value = SCL_F
        sheet.cell(id_row, 43).value = SDA_R
        sheet.cell(id_row, 47).value = SDA_F

        sheet.cell(id_row, 50).value = tsu_mea
        sheet.cell(id_row, 53).value = thd_mea

        sheet.cell(id_row, 56).value = sta_mea
        sheet.cell(id_row, 59).value = sto_mea

    if int(num_csv) == 1:

        # input the voltage value to the test plan
        sheet.cell(id_row, 11).value = SCL_Max
        sheet.cell(id_row, 14).value = SCL_Min
        sheet.cell(id_row, 18).value = SCL_H
        sheet.cell(id_row, 22).value = SCL_L

        sheet.cell(id_row + 1, 11).value = SDA_Max
        sheet.cell(id_row + 1, 14).value = SDA_Min
        sheet.cell(id_row + 1, 18).value = SDA_H
        sheet.cell(id_row + 1, 22).value = SDA_L

    if int(num_csv) == 6:

        sheet.cell(id_row + 1, 11).value = SDA_Max
        sheet.cell(id_row + 1, 14).value = SDA_Min
        sheet.cell(id_row + 1, 18).value = SDA_H
        sheet.cell(id_row + 1, 22).value = SDA_L

        sheet.cell(id_row, 43).value = SDA_R
        sheet.cell(id_row, 47).value = SDA_F

        sheet.cell(id_row, 50).value = tsu_mea
        sheet.cell(id_row, 53).value = thd_mea

        sheet.cell(id_row, 56).value = sta_mea
        sheet.cell(id_row, 59).value = sto_mea

    # margin calculation
    if int(num_csv) == 7:
        # SCL&SDA_MAX margin
        if type(sheet.cell(id_row, 10).value) is not str:
            # print(type(sheet.cell(id_row, 10).value))
            sheet.cell(id_row, 12).value = (sheet.cell(id_row, 10).value - sheet.cell(id_row, 11).value) / sheet.cell(id_row, 10).value
            sheet.cell(id_row + 1, 12).value = (sheet.cell(id_row, 10).value - sheet.cell(id_row + 1, 11).value) / \
                                               sheet.cell(id_row, 10).value
            # set the margin value format
            sheet.cell(id_row, 12).number_format = '0%'
            sheet.cell(id_row + 1, 12).number_format = '0%'
            # set the cell color
            if sheet.cell(id_row, 12).value >= 0.05:
                sheet.cell(id_row, 12).font = font_blue
                sheet.cell(id_row, 11).font = font_blue
            if 0 < sheet.cell(id_row, 12).value < 0.05:
                sheet.cell(id_row, 12).fill = fill_yellow
                sheet.cell(id_row, 11).fill = fill_yellow
            if sheet.cell(id_row, 12).value <= 0:
                sheet.cell(id_row, 12).fill = fill_red
                sheet.cell(id_row, 11).fill = fill_red

            if sheet.cell(id_row + 1, 12).value >= 0.05:
                sheet.cell(id_row + 1, 12).font = font_blue
                sheet.cell(id_row + 1, 11).font = font_blue
            if 0 < sheet.cell(id_row + 1, 12).value < 0.05:
                sheet.cell(id_row + 1, 12).fill = fill_yellow
                sheet.cell(id_row + 1, 11).fill = fill_yellow
            if sheet.cell(id_row + 1, 12).value <= 0:
                sheet.cell(id_row + 1, 12).fill = fill_red
                sheet.cell(id_row + 1, 11).fill = fill_red

        elif sheet.cell(id_row, 10).value == 'NA':
            sheet.cell(id_row, 12).value = '/'
            sheet.cell(id_row + 1, 12).value = '/'
            # set the cell color
            sheet.cell(id_row, 12).font = font_blue
            sheet.cell(id_row, 11).font = font_blue
            sheet.cell(id_row + 1, 12).font = font_blue
            sheet.cell(id_row + 1, 11).font = font_blue

        # SCL&SDA_Min margin
        if type(sheet.cell(id_row, 13).value) is not str:
            sheet.cell(id_row, 15).value = (sheet.cell(id_row, 13).value - sheet.cell(id_row, 14).value) / sheet.cell(
            id_row, 13).value
            sheet.cell(id_row + 1, 15).value = (sheet.cell(id_row, 13).value - sheet.cell(id_row + 1, 14).value) / \
                                               sheet.cell(id_row, 13).value
            sheet.cell(id_row, 15).number_format = '0%'
            sheet.cell(id_row + 1, 15).number_format = '0%'
            # set the cell color
            if sheet.cell(id_row, 15).value >= 0.05:
                sheet.cell(id_row, 15).font = font_blue
                sheet.cell(id_row, 14).font = font_blue
            if 0 < sheet.cell(id_row, 15).value < 0.05:
                sheet.cell(id_row, 15).fill = fill_yellow
                sheet.cell(id_row, 14).fill = fill_yellow
            if sheet.cell(id_row, 15).value <= 0:
                sheet.cell(id_row, 15).fill = fill_red
                sheet.cell(id_row, 14).fill = fill_red

            if sheet.cell(id_row + 1, 15).value >= 0.05:
                sheet.cell(id_row + 1, 15).font = font_blue
                sheet.cell(id_row + 1, 14).font = font_blue
            if 0 < sheet.cell(id_row + 1, 15).value < 0.05:
                sheet.cell(id_row + 1, 15).fill = fill_yellow
                sheet.cell(id_row + 1, 14).fill = fill_yellow
            if sheet.cell(id_row + 1, 15).value <= 0:
                sheet.cell(id_row + 1, 15).fill = fill_red
                sheet.cell(id_row + 1, 14).fill = fill_red

        elif sheet.cell(id_row, 13).value == 'NA':
            sheet.cell(id_row, 15).value = '/'
            sheet.cell(id_row + 1, 15).value = '/'
            sheet.cell(id_row, 15).font = font_blue
            sheet.cell(id_row, 14).font = font_blue
            sheet.cell(id_row + 1, 15).font = font_blue
            sheet.cell(id_row + 1, 14).font = font_blue

        # # SCL_H&SDA_H margin
        if type(sheet.cell(id_row, 16).value) is int or float:
            sheet.cell(id_row, 19).value = (sheet.cell(id_row, 18).value - sheet.cell(id_row, 16).value) / sheet.cell(
            id_row, 16).value
            sheet.cell(id_row + 1, 19).value = (sheet.cell(id_row + 1, 18).value - sheet.cell(id_row, 16).value) / \
                                               sheet.cell(id_row, 16).value
            sheet.cell(id_row, 19).number_format = '0%'
            sheet.cell(id_row + 1, 19).number_format = '0%'
            # set the cell color
            if sheet.cell(id_row, 19).value >= 0.05:
                sheet.cell(id_row, 19).font = font_blue
                sheet.cell(id_row, 18).font = font_blue
            if 0 < sheet.cell(id_row, 19).value < 0.05:
                sheet.cell(id_row, 19).fill = fill_yellow
                sheet.cell(id_row, 18).fill = fill_yellow
            if sheet.cell(id_row, 19).value <= 0:
                sheet.cell(id_row, 19).fill = fill_red
                sheet.cell(id_row, 18).fill = fill_red

            if sheet.cell(id_row + 1, 19).value >= 0.05:
                sheet.cell(id_row + 1, 19).font = font_blue
                sheet.cell(id_row + 1, 18).font = font_blue
            if 0 < sheet.cell(id_row + 1, 19).value < 0.05:
                sheet.cell(id_row + 1, 19).fill = fill_yellow
                sheet.cell(id_row + 1, 18).fill = fill_yellow
            if sheet.cell(id_row + 1, 19).value <= 0:
                sheet.cell(id_row + 1, 19).fill = fill_red
                sheet.cell(id_row + 1, 18).fill = fill_red

        elif sheet.cell(id_row, 16).value == 'NA':
            sheet.cell(id_row, 19).value = '/'
            sheet.cell(id_row + 1, 19).value = '/'
            sheet.cell(id_row, 19).font = font_blue
            sheet.cell(id_row + 1, 19).font = font_blue

        # # SCL_L&SDA_L margin
        if type(sheet.cell(id_row, 21).value) is int or float:
            sheet.cell(id_row, 23).value = (sheet.cell(id_row, 21).value - sheet.cell(id_row, 22).value) / sheet.cell(
            id_row, 21).value
            sheet.cell(id_row + 1, 23).value = (sheet.cell(id_row, 21).value - sheet.cell(id_row + 1, 22).value) / \
                                               sheet.cell(id_row, 21).value
            sheet.cell(id_row, 23).number_format = '0%'
            sheet.cell(id_row + 1, 23).number_format = '0%'
            # set the cell color
            if sheet.cell(id_row, 23).value >= 0.05:
                sheet.cell(id_row, 23).font = font_blue
                sheet.cell(id_row, 22).font = font_blue
            if 0 < sheet.cell(id_row, 23).value < 0.05:
                sheet.cell(id_row, 23).fill = fill_yellow
                sheet.cell(id_row, 22).fill = fill_yellow
            if sheet.cell(id_row, 23).value <= 0:
                sheet.cell(id_row, 23).fill = fill_red
                sheet.cell(id_row, 22).fill = fill_red

            if sheet.cell(id_row + 1, 23).value >= 0.05:
                sheet.cell(id_row + 1, 23).font = font_blue
                sheet.cell(id_row + 1, 22).font = font_blue
            if 0 < sheet.cell(id_row + 1, 23).value < 0.05:
                sheet.cell(id_row + 1, 23).fill = fill_yellow
                sheet.cell(id_row + 1, 22).fill = fill_yellow
            if sheet.cell(id_row + 1, 23).value <= 0:
                sheet.cell(id_row + 1, 23).fill = fill_red
                sheet.cell(id_row + 1, 22).fill = fill_red

        elif sheet.cell(id_row, 21).value == 'NA':
            sheet.cell(id_row, 23).value = '/'
            sheet.cell(id_row + 1, 23).value = '/'
            # set the cell color
            sheet.cell(id_row, 23).font = font_blue
            sheet.cell(id_row + 1, 23).font = font_blue

        # frequency margin
        if type(sheet.cell(id_row, 24).value) is not str:
            if sheet.cell(id_row, 25).value < sheet.cell(id_row, 24).value:
                sheet.cell(id_row, 26).value = '/'
                sheet.cell(id_row, 26).font = font_blue
                sheet.cell(id_row, 25).font = font_blue

        # high time margin
        if type(sheet.cell(id_row, 27).value) is not str:
            sheet.cell(id_row, 29).value = (sheet.cell(id_row, 28).value - sheet.cell(id_row, 27).value) / sheet.cell(
                id_row, 27).value
            sheet.cell(id_row, 29).number_format = '0%'
            # set the cell color
            if sheet.cell(id_row, 29).value >= 0.05:
                sheet.cell(id_row, 29).font = font_blue
                sheet.cell(id_row, 28).font = font_blue
            if 0 < sheet.cell(id_row, 29).value < 0.05:
                sheet.cell(id_row, 29).fill = fill_yellow
                sheet.cell(id_row, 28).fill = fill_yellow
            if sheet.cell(id_row, 29).value <= 0:
                sheet.cell(id_row, 29).fill = fill_red
                sheet.cell(id_row, 28).fill = fill_red

        # low time margin
        if type(sheet.cell(id_row, 30).value) is not str:
            sheet.cell(id_row, 32).value = (sheet.cell(id_row, 31).value - sheet.cell(id_row, 30).value) / sheet.cell(
                id_row, 30).value
            sheet.cell(id_row, 32).number_format = '0%'
            # set the cell color
            if sheet.cell(id_row, 32).value >= 0.05:
                sheet.cell(id_row, 32).font = font_blue
                sheet.cell(id_row, 31).font = font_blue
            if 0 < sheet.cell(id_row, 32).value < 0.05:
                sheet.cell(id_row, 32).fill = fill_yellow
                sheet.cell(id_row, 31).fill = fill_yellow
            if sheet.cell(id_row, 32).value <= 0:
                sheet.cell(id_row, 32).fill = fill_red
                sheet.cell(id_row, 31).fill = fill_red

        # SCL_Rise margin
        if type(sheet.cell(id_row, 34).value) is not str:
            if sheet.cell(id_row, 33).value < sheet.cell(id_row, 35).value < sheet.cell(id_row, 34).value:
                min_SCL_R = sheet.cell(id_row, 35).value - sheet.cell(id_row, 33).value
                max_SCL_R = sheet.cell(id_row, 34).value - sheet.cell(id_row, 35).value
                if max_SCL_R < min_SCL_R:
                    sheet.cell(id_row, 36).value = max_SCL_R / sheet.cell(id_row, 34).value
                else:
                    sheet.cell(id_row, 36).value = min_SCL_R / sheet.cell(id_row, 33).value
                sheet.cell(id_row, 36).number_format = '0%'
                # set the cell color
                if sheet.cell(id_row, 36).value >= 0.05:
                    sheet.cell(id_row, 36).font = font_blue
                    sheet.cell(id_row, 35).font = font_blue
                if 0 < sheet.cell(id_row, 36).value < 0.05:
                    sheet.cell(id_row, 36).fill = fill_yellow
                    sheet.cell(id_row, 35).fill = fill_yellow
            else:
                sheet.cell(id_row, 36).fill = fill_red
                sheet.cell(id_row, 35).fill = fill_red

        # SCL_Fall margin
        if type(sheet.cell(id_row, 38).value) is not str:
            sheet.cell(id_row, 40).value = (sheet.cell(id_row, 38).value - sheet.cell(id_row, 39).value) / sheet.cell(
                id_row, 38).value
            sheet.cell(id_row, 40).number_format = '0%'
            # set the cell color
            if sheet.cell(id_row, 40).value >= 0.05:
                sheet.cell(id_row, 40).font = font_blue
                sheet.cell(id_row, 39).font = font_blue
            if 0 < sheet.cell(id_row, 40).value < 0.05:
                sheet.cell(id_row, 40).fill = fill_yellow
                sheet.cell(id_row, 39).fill = fill_yellow
            if sheet.cell(id_row, 40).value <= 0:
                sheet.cell(id_row, 40).fill = fill_red
                sheet.cell(id_row, 39).fill = fill_red

        # SDA_Rise margin
        if type(sheet.cell(id_row, 42).value) is not str:
            if sheet.cell(id_row, 41).value < sheet.cell(id_row, 43).value < sheet.cell(id_row, 42).value:
                min_SDA_R = sheet.cell(id_row, 43).value - sheet.cell(id_row, 41).value
                max_SDA_R = sheet.cell(id_row, 42).value - sheet.cell(id_row, 43).value
                if max_SDA_R < min_SDA_R:
                    sheet.cell(id_row, 44).value = max_SDA_R / sheet.cell(id_row, 42).value
                else:
                    sheet.cell(id_row, 44).value = min_SDA_R / sheet.cell(id_row, 41).value
                sheet.cell(id_row, 44).number_format = '0%'
                # set the cell color
                if sheet.cell(id_row, 44).value >= 0.05:
                    sheet.cell(id_row, 44).font = font_blue
                    sheet.cell(id_row, 43).font = font_blue
                if 0 < sheet.cell(id_row, 44).value < 0.05:
                    sheet.cell(id_row, 44).fill = fill_yellow
                    sheet.cell(id_row, 43).fill = fill_yellow
            else:
                sheet.cell(id_row, 44).fill = fill_red
                sheet.cell(id_row, 43).fill = fill_red

        # SDA_Fall margin
        if type(sheet.cell(id_row, 46).value) is not str:
            min_SDA_F = sheet.cell(id_row, 47).value - sheet.cell(id_row, 45).value
            max_SDA_F = sheet.cell(id_row, 46).value - sheet.cell(id_row, 47).value
            if min_SDA_F < max_SDA_F:
                sheet.cell(id_row, 48).value = min_SDA_F / sheet.cell(id_row, 46).value
            else:
                sheet.cell(id_row, 48).value = max_SDA_F / sheet.cell(id_row, 46).value
            sheet.cell(id_row, 48).number_format = '0%'
            # set the cell color
            if sheet.cell(id_row, 48).value >= 0.05:
                sheet.cell(id_row, 48).font = font_blue
                sheet.cell(id_row, 47).font = font_blue
            if 0 < sheet.cell(id_row, 48).value < 0.05:
                sheet.cell(id_row, 48).fill = fill_yellow
                sheet.cell(id_row, 47).fill = fill_yellow
            if sheet.cell(id_row, 48).value <= 0:
                sheet.cell(id_row, 48).fill = fill_red
                sheet.cell(id_row, 47).fill = fill_red

        # Tsu margin
        if type(sheet.cell(id_row, 49).value) is int or float:
            sheet.cell(id_row, 51).value = (sheet.cell(id_row, 50).value - sheet.cell(id_row, 49).value) / sheet.cell(
            id_row, 49).value
            sheet.cell(id_row, 51).number_format = '0%'
            # set the cell color
            if sheet.cell(id_row, 51).value >= 0.05:
                sheet.cell(id_row, 51).font = font_blue
                sheet.cell(id_row, 50).font = font_blue
            if 0 < sheet.cell(id_row, 51).value < 0.05:
                sheet.cell(id_row, 51).fill = fill_yellow
                sheet.cell(id_row, 50).fill = fill_yellow
            if sheet.cell(id_row, 51).value <= 0:
                sheet.cell(id_row, 51).fill = fill_red
                sheet.cell(id_row, 50).fill = fill_red

        # Thd margin
        if sheet.cell(id_row, 52).value == 0:
            if sheet.cell(id_row, 53).value > sheet.cell(id_row, 52).value:
                sheet.cell(id_row, 54).value = '/'
                sheet.cell(id_row, 54).font = font_blue
                sheet.cell(id_row, 53).font = font_blue

        # Sta margin
        if type(sheet.cell(id_row, 55).value) is int or float:
            if sheet.cell(id_row, 56).value > sheet.cell(id_row, 55).value:
                sheet.cell(id_row, 57).value = (sheet.cell(id_row, 56).value - sheet.cell(id_row, 55).value) / sheet.cell(id_row, 55).value
                sheet.cell(id_row, 57).number_format = '0%'
                # set the cell color
                if sheet.cell(id_row, 57).value >= 0.05:
                    sheet.cell(id_row, 57).font = font_blue
                    sheet.cell(id_row, 56).font = font_blue
                    # sheet.cell(id_row, 8).value = 'Pass'
                if 0 < sheet.cell(id_row, 57).value < 0.05:
                    sheet.cell(id_row, 57).fill = fill_yellow
                    sheet.cell(id_row, 56).fill = fill_yellow
                    # sheet.cell(id_row, 8).value = 'Marginal Pass'
            else:
                sheet.cell(id_row, 57).fill = fill_red
                sheet.cell(id_row, 56).fill = fill_red
                # sheet.cell(id_row, 8).value = 'Fail'

        # Sto margin
        if type(sheet.cell(id_row, 58).value) is int or float:
            if sheet.cell(id_row, 59).value > sheet.cell(id_row, 58).value:
                sheet.cell(id_row, 60).value = (sheet.cell(id_row, 59).value - sheet.cell(id_row, 58).value) / sheet.cell(id_row, 58).value
                sheet.cell(id_row, 60).number_format = '0%'
                # set the cell color
                if sheet.cell(id_row, 60).value >= 0.05:
                    sheet.cell(id_row, 60).font = font_blue
                    sheet.cell(id_row, 59).font = font_blue
                    # sheet.cell(id_row, 8).value = 'Pass'
                if 0 < sheet.cell(id_row, 60).value < 0.05:
                    sheet.cell(id_row, 60).fill = fill_yellow
                    sheet.cell(id_row, 59).fill = fill_yellow
                    # sheet.cell(id_row, 8).value = 'Marginal Pass'
            else:
                sheet.cell(id_row, 60).fill = fill_red
                sheet.cell(id_row, 59).fill = fill_red
                # sheet.cell(id_row, 8).value = 'Fail'

    # margin calculation
    if int(num_csv) == 1:
        # SCL_MAX&SDA_MAX margin
        if type(sheet.cell(id_row, 10).value) is int or float:
            sheet.cell(id_row, 12).value = (sheet.cell(id_row, 10).value - sheet.cell(id_row, 11).value) / sheet.cell(
            id_row, 10).value
            sheet.cell(id_row + 1, 12).value = (sheet.cell(id_row, 10).value - sheet.cell(id_row + 1, 11).value) / \
                                               sheet.cell(id_row, 10).value
            sheet.cell(id_row, 12).number_format = '0%'
            sheet.cell(id_row + 1, 12).number_format = '0%'
            # set the cell color
            if sheet.cell(id_row, 12).value >= 0.05:
                sheet.cell(id_row, 12).font = font_blue
                sheet.cell(id_row, 11).font = font_blue
            if 0 < sheet.cell(id_row, 12).value < 0.05:
                sheet.cell(id_row, 12).fill = fill_yellow
                sheet.cell(id_row, 11).fill = fill_yellow
            if sheet.cell(id_row, 12).value <= 0:
                sheet.cell(id_row, 12).fill = fill_red
                sheet.cell(id_row, 11).fill = fill_red

            if sheet.cell(id_row + 1, 12).value >= 0.05:
                sheet.cell(id_row + 1, 12).font = font_blue
                sheet.cell(id_row + 1, 11).font = font_blue
            if 0 < sheet.cell(id_row + 1, 12).value < 0.05:
                sheet.cell(id_row + 1, 12).fill = fill_yellow
                sheet.cell(id_row + 1, 11).fill = fill_yellow
            if sheet.cell(id_row + 1, 12).value <= 0:
                sheet.cell(id_row + 1, 12).fill = fill_red
                sheet.cell(id_row + 1, 11).fill = fill_red

        elif sheet.cell(id_row, 10).value == 'NA':
            sheet.cell(id_row, 12).value = '/'
            sheet.cell(id_row + 1, 12).value = '/'
            sheet.cell(id_row, 12).font = font_blue
            sheet.cell(id_row + 1, 12).font = font_blue

        # SCL_Min&SDA_Min margin
        if type(sheet.cell(id_row, 13).value) is int or float:
            sheet.cell(id_row, 15).value = (sheet.cell(id_row, 13).value - sheet.cell(id_row, 14).value) / sheet.cell(
            id_row, 13).value
            sheet.cell(id_row + 1, 15).value = (sheet.cell(id_row, 13).value - sheet.cell(id_row + 1, 14).value) / \
                                               sheet.cell(id_row, 13).value
            sheet.cell(id_row, 15).number_format = '0%'
            sheet.cell(id_row + 1, 15).number_format = '0%'
            # set the cell color
            if sheet.cell(id_row, 15).value >= 0.05:
                sheet.cell(id_row, 15).font = font_blue
                sheet.cell(id_row, 14).font = font_blue
            if 0 < sheet.cell(id_row, 15).value < 0.05:
                sheet.cell(id_row, 15).fill = fill_yellow
                sheet.cell(id_row, 14).fill = fill_yellow
            if sheet.cell(id_row, 15).value <= 0:
                sheet.cell(id_row, 15).fill = fill_red
                sheet.cell(id_row, 14).fill = fill_red

            if sheet.cell(id_row + 1, 15).value >= 0.05:
                sheet.cell(id_row + 1, 15).font = font_blue
                sheet.cell(id_row + 1, 14).font = font_blue
            if 0 < sheet.cell(id_row + 1, 15).value < 0.05:
                sheet.cell(id_row + 1, 15).fill = fill_yellow
                sheet.cell(id_row + 1, 14).fill = fill_yellow
            if sheet.cell(id_row + 1, 15).value <= 0:
                sheet.cell(id_row + 1, 15).fill = fill_red
                sheet.cell(id_row + 1, 14).fill = fill_red

        elif sheet.cell(id_row, 13).value == 'NA':
            sheet.cell(id_row, 15).value = '/'
            sheet.cell(id_row + 1, 15).value = '/'
            sheet.cell(id_row, 15).font = font_blue
            sheet.cell(id_row + 1, 15).font = font_blue

        # SCL_H margin
        # if type(sheet.cell(id_row, 16).value) is int or float:
        #     sheet.cell(id_row, 19).value = (sheet.cell(id_row, 18).value - sheet.cell(id_row, 16).value) / sheet.cell(
        #         id_row, 16).value
        # elif sheet.cell(id_row, 16).value == 'NA':
        #     sheet.cell(id_row, 19).value = '/'
        sheet.cell(id_row, 19).value = '/'
        sheet.cell(id_row, 19).font = font_blue
        sheet.cell(id_row, 18).font = font_blue

        # SDA_H margin
        if type(sheet.cell(id_row + 1, 16).value) is int or float:
            sheet.cell(id_row + 1, 19).value = (sheet.cell(id_row + 1, 18).value - sheet.cell(id_row + 1, 16).value) / \
                                               sheet.cell(id_row + 1, 16).value
            sheet.cell(id_row + 1, 19).number_format = '0%'
            # set the cell color
            if sheet.cell(id_row + 1, 19).value >= 0.05:
                sheet.cell(id_row + 1, 19).font = font_blue
                sheet.cell(id_row + 1, 18).font = font_blue
            if 0 < sheet.cell(id_row + 1, 19).value < 0.05:
                sheet.cell(id_row + 1, 19).fill = fill_yellow
                sheet.cell(id_row + 1, 18).fill = fill_yellow
            if sheet.cell(id_row + 1, 19).value <= 0:
                sheet.cell(id_row + 1, 19).fill = fill_red
                sheet.cell(id_row + 1, 18).fill = fill_red

        elif sheet.cell(id_row + 1, 16).value == 'NA':
            sheet.cell(id_row + 1, 19).value = '/'
            sheet.cell(id_row + 1, 19).font = font_blue
            sheet.cell(id_row + 1, 18).font = font_blue

        # SCL_L margin
        if type(sheet.cell(id_row, 21).value) is int or float:
            if sheet.cell(id_row, 22).value < sheet.cell(id_row, 21).value:
                # min_SCL_L = sheet.cell(id_row, 22).value - sheet.cell(id_row, 20).value
                max_SCL_L = sheet.cell(id_row, 21).value - sheet.cell(id_row, 22).value
                # if min_SCL_L < max_SCL_L:
                #     sheet.cell(id_row, 23).value = min_SCL_L / sheet.cell(id_row, 20).value
                # else:
                sheet.cell(id_row, 23).value = max_SCL_L / sheet.cell(id_row, 21).value
            sheet.cell(id_row, 23).number_format = '0%'
            # set the cell color
            if sheet.cell(id_row, 23).value >= 0.05:
                sheet.cell(id_row, 23).font = font_blue
                sheet.cell(id_row, 22).font = font_blue
            if 0 < sheet.cell(id_row, 23).value < 0.05:
                sheet.cell(id_row, 23).fill = fill_yellow
                sheet.cell(id_row, 22).fill = fill_yellow
            if sheet.cell(id_row, 23).value <= 0:
                sheet.cell(id_row, 23).fill = fill_red
                sheet.cell(id_row, 22).fill = fill_red

        elif sheet.cell(id_row, 21).value == 'NA':
            sheet.cell(id_row, 23).value = '/'
            sheet.cell(id_row, 23).font = font_blue

        # SDA_L margin
        if type(sheet.cell(id_row + 1, 21).value) is int or float:
            sheet.cell(id_row + 1, 23).value = (sheet.cell(id_row + 1, 21).value - sheet.cell(id_row + 1, 22).value) / \
                                               sheet.cell(id_row + 1, 21).value
            sheet.cell(id_row + 1, 23).number_format = '0%'
            # set the cell color
            if sheet.cell(id_row + 1, 23).value >= 0.05:
                sheet.cell(id_row + 1, 23).font = font_blue
                sheet.cell(id_row + 1, 22).font = font_blue
            if 0 < sheet.cell(id_row + 1, 23).value < 0.05:
                sheet.cell(id_row + 1, 23).fill = fill_yellow
                sheet.cell(id_row + 1, 22).fill = fill_yellow
            if sheet.cell(id_row + 1, 23).value <= 0:
                sheet.cell(id_row + 1, 23).fill = fill_red
                sheet.cell(id_row + 1, 22).fill = fill_red

        elif sheet.cell(id_row + 1, 21).value == 'NA':
            sheet.cell(id_row + 1, 23).value = '/'
            sheet.cell(id_row + 1, 23).font = font_blue

    # margin calculation
    if int(num_csv) == 6:
        # SDA_Max margin
        if type(sheet.cell(id_row + 1, 10).value) is int or float:
            sheet.cell(id_row + 1, 12).value = (sheet.cell(id_row + 1, 10).value - sheet.cell(id_row + 1, 11).value)/\
                                               sheet.cell(id_row + 1, 10).value
            sheet.cell(id_row + 1, 12).number_format = '0%'
            # set the cell color
            if sheet.cell(id_row + 1, 12).value >= 0.05:
                sheet.cell(id_row + 1, 12).font = font_blue
                sheet.cell(id_row + 1, 11).font = font_blue
            if 0 < sheet.cell(id_row + 1, 12).value < 0.05:
                sheet.cell(id_row + 1, 12).fill = fill_yellow
                sheet.cell(id_row + 1, 11).fill = fill_yellow
            if sheet.cell(id_row + 1, 12).value <= 0:
                sheet.cell(id_row + 1, 12).fill = fill_red
                sheet.cell(id_row + 1, 11).fill = fill_red

        elif sheet.cell(id_row + 1, 10).value == 'NA':
            sheet.cell(id_row + 1, 12).value = '/'
            sheet.cell(id_row + 1, 12).font = font_blue
            sheet.cell(id_row + 1, 11).font = font_blue

        # SDA_Min margin
        if type(sheet.cell(id_row + 1, 13).value) is int or float:
            sheet.cell(id_row + 1, 15).value = (sheet.cell(id_row + 1, 13).value - sheet.cell(id_row + 1, 14).value) / \
                                               sheet.cell(id_row + 1, 13).value
            sheet.cell(id_row + 1, 15).number_format = '0%'
            # set the cell color
            if sheet.cell(id_row + 1, 15).value >= 0.05:
                sheet.cell(id_row + 1, 15).font = font_blue
                sheet.cell(id_row + 1, 14).font = font_blue
            if 0 < sheet.cell(id_row + 1, 15).value < 0.05:
                sheet.cell(id_row + 1, 15).fill = fill_yellow
                sheet.cell(id_row + 1, 14).fill = fill_yellow
            if sheet.cell(id_row + 1, 15).value <= 0:
                sheet.cell(id_row + 1, 15).fill = fill_red
                sheet.cell(id_row + 1, 14).fill = fill_red

        elif sheet.cell(id_row + 1, 13).value == 'NA':
            sheet.cell(id_row + 1, 15).value = '/'
            sheet.cell(id_row + 1, 15).font = font_blue
            sheet.cell(id_row + 1, 14).font = font_blue

        # SDA_H margin
        if type(sheet.cell(id_row + 1, 16).value) is int or float:
            sheet.cell(id_row + 1, 19).value = (sheet.cell(id_row + 1, 18).value - sheet.cell(id_row + 1, 16).value) / \
                                               sheet.cell(id_row + 1, 16).value
            sheet.cell(id_row + 1, 19).number_format = '0%'
            # set the cell color
            if sheet.cell(id_row + 1, 19).value >= 0.05:
                sheet.cell(id_row + 1, 19).font = font_blue
                sheet.cell(id_row + 1, 18).font = font_blue
            if 0 < sheet.cell(id_row + 1, 19).value < 0.05:
                sheet.cell(id_row + 1, 19).fill = fill_yellow
                sheet.cell(id_row + 1, 18).fill = fill_yellow
            if sheet.cell(id_row + 1, 19).value <= 0:
                sheet.cell(id_row + 1, 19).fill = fill_red
                sheet.cell(id_row + 1, 18).fill = fill_red

        elif sheet.cell(id_row + 1, 16).value == 'NA':
            sheet.cell(id_row + 1, 19).value = '/'
            sheet.cell(id_row + 1, 19).font = font_blue
            sheet.cell(id_row + 1, 18).font = font_blue

        # SDA_L margin
        if type(sheet.cell(id_row + 1, 21).value) is int or float:
            sheet.cell(id_row + 1, 23).value = (sheet.cell(id_row + 1, 21).value - sheet.cell(id_row + 1, 22).value) / \
                                               sheet.cell(id_row + 1, 21).value
            sheet.cell(id_row + 1, 23).number_format = '0%'
            # set the cell color
            if sheet.cell(id_row + 1, 23).value >= 0.05:
                sheet.cell(id_row + 1, 23).font = font_blue
                sheet.cell(id_row + 1, 22).font = font_blue
            if 0 < sheet.cell(id_row + 1, 23).value < 0.05:
                sheet.cell(id_row + 1, 23).fill = fill_yellow
                sheet.cell(id_row + 1, 22).fill = fill_yellow
            if sheet.cell(id_row + 1, 23).value <= 0:
                sheet.cell(id_row + 1, 23).fill = fill_red
                sheet.cell(id_row + 1, 22).fill = fill_red

        elif sheet.cell(id_row + 1, 21).value == 'NA':
            sheet.cell(id_row + 1, 23).value = '/'
            sheet.cell(id_row + 1, 23).font = font_blue
            sheet.cell(id_row + 1, 22).font = font_blue

        # SDA_R margin
        if type(sheet.cell(id_row, 42).value) is int or float:  # SDA_Rise_max spec is a data
            if sheet.cell(id_row, 41).value < sheet.cell(id_row, 43).value < sheet.cell(id_row, 42).value:
                min_SDA_R = sheet.cell(id_row, 43).value - sheet.cell(id_row, 41).value
                max_SDA_R = sheet.cell(id_row, 42).value - sheet.cell(id_row, 43).value
                if max_SDA_R < min_SDA_R:
                    sheet.cell(id_row, 44).value = max_SDA_R / sheet.cell(id_row, 42).value
                else:
                    sheet.cell(id_row, 44).value = min_SDA_R / sheet.cell(id_row, 41).value
                sheet.cell(id_row, 44).number_format = '0%'
                # set the cell color
                if sheet.cell(id_row, 44).value >= 0.05:
                    sheet.cell(id_row, 44).font = font_blue
                    sheet.cell(id_row, 43).font = font_blue
                if 0 < sheet.cell(id_row, 44).value < 0.05:
                    sheet.cell(id_row, 44).fill = fill_yellow
                    sheet.cell(id_row, 43).fill = fill_yellow
            else:
                sheet.cell(id_row, 44).value = '/'
                sheet.cell(id_row, 44).fill = fill_red
                sheet.cell(id_row, 43).fill = fill_red

        # SDA_F margin
        if type(sheet.cell(id_row, 46).value) is int or float:
            min_SDA_F = sheet.cell(id_row, 47).value - sheet.cell(id_row, 45).value
            max_SDA_F = sheet.cell(id_row, 46).value - sheet.cell(id_row, 47).value
            if min_SDA_F < max_SDA_F:
                sheet.cell(id_row, 48).value = min_SDA_F / sheet.cell(id_row, 46).value
            else:
                sheet.cell(id_row, 48).value = max_SDA_F / sheet.cell(id_row, 46).value
            sheet.cell(id_row, 48).number_format = '0%'
            # set the cell color
            if sheet.cell(id_row, 48).value >= 0.05:
                sheet.cell(id_row, 48).font = font_blue
                sheet.cell(id_row, 47).font = font_blue
            if 0 < sheet.cell(id_row, 48).value < 0.05:
                sheet.cell(id_row, 48).fill = fill_yellow
                sheet.cell(id_row, 47).fill = fill_yellow
            if sheet.cell(id_row, 48).value <= 0:
                sheet.cell(id_row, 48).fill = fill_red
                sheet.cell(id_row, 47).fill = fill_red

        # Tsu margin
        if type(sheet.cell(id_row, 49).value) is int or float:
            sheet.cell(id_row, 51).value = (sheet.cell(id_row, 50).value - sheet.cell(id_row, 49).value) / \
                                           sheet.cell(id_row, 49).value
            sheet.cell(id_row, 51).number_format = '0%'
            # set the cell color
            if sheet.cell(id_row, 51).value >= 0.05:
                sheet.cell(id_row, 51).font = font_blue
                sheet.cell(id_row, 50).font = font_blue
            if 0 < sheet.cell(id_row, 51).value < 0.05:
                sheet.cell(id_row, 51).fill = fill_yellow
                sheet.cell(id_row, 50).fill = fill_yellow
            if sheet.cell(id_row, 51).value <= 0:
                sheet.cell(id_row, 51).fill = fill_red
                sheet.cell(id_row, 50).fill = fill_red

        # Thd margin
        if sheet.cell(id_row, 52).value == 0:
            if sheet.cell(id_row, 53).value > sheet.cell(id_row, 52).value:
                sheet.cell(id_row, 54).value = '/'
                sheet.cell(id_row, 54).font = font_blue
                sheet.cell(id_row, 53).font = font_blue

        # Sta margin
        if type(sheet.cell(id_row, 55).value) is int or float:
            sheet.cell(id_row, 57).value = (sheet.cell(id_row, 56).value - sheet.cell(id_row, 55).value) / \
                                           sheet.cell(id_row, 55).value
            sheet.cell(id_row, 57).number_format = '0%'
            # set the cell color
            if sheet.cell(id_row, 57).value >= 0.05:
                sheet.cell(id_row, 57).font = font_blue
                sheet.cell(id_row, 56).font = font_blue
            if 0 < sheet.cell(id_row, 57).value < 0.05:
                sheet.cell(id_row, 57).fill = fill_yellow
                sheet.cell(id_row, 56).fill = fill_yellow
            if sheet.cell(id_row, 57).value <= 0:
                sheet.cell(id_row, 57).fill = fill_red
                sheet.cell(id_row, 56).fill = fill_red

        # Sto margin
        if type(sheet.cell(id_row, 58).value) is int or float:
            sheet.cell(id_row, 60).value = (sheet.cell(id_row, 59).value - sheet.cell(id_row, 58).value) / \
                                           sheet.cell(id_row, 58).value
            sheet.cell(id_row, 60).number_format = '0%'
            # set the cell color
            if sheet.cell(id_row, 60).value >= 0.05:
                sheet.cell(id_row, 60).font = font_blue
                sheet.cell(id_row, 59).font = font_blue
            if 0 < sheet.cell(id_row, 60).value < 0.05:
                sheet.cell(id_row, 60).fill = fill_yellow
                sheet.cell(id_row, 59).fill = fill_yellow
            if sheet.cell(id_row, 60).value <= 0:
                sheet.cell(id_row, 60).fill = fill_red
                sheet.cell(id_row, 59).fill = fill_red

    # format the report

wb.save(planPath)
