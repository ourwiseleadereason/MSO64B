# data: 2023-10-17

import os
import csv
import pandas as pd
import openpyxl


class PowerNoise:
    pass

    def __init__(self):
        pass


# handle csv files to get the measured data
def csv_handle(path, target_row, target_col):

    largest_column_count = 0
    # target_row = 'RMS'
    row_target = 0
    # target_col = 'Mean\''
    col_target = 0

    with open(path, 'r') as f:
        # reader = csv.reader(f)
        lines = f.readlines()
        # print(lines)
        for line in lines:
            column_count = len(line.split(',')) + 1
            largest_column_count = column_count if largest_column_count < column_count else largest_column_count
    f.close()

    column_names = [i for i in range(0, largest_column_count)]

    df = pd.read_csv(path, header=None, delimiter=',', names=column_names)
    row_count = df.shape[0]
    col_count = df.shape[1]
    print(row_count, col_count)

    for row in range(0, int(row_count)):
        for col in range(0, int(col_count)):

            if df[col][row] == target_row:
                row_target = row
                # print(row_target)
            if df[col][row] == target_col:
                col_target = col
    # print(row_target, col_target)
    mea_data = df[col_target][row_target]
    return mea_data


def report_handle(file, row, col):

    # wb = openpyxl.load_workbook(r'E:\Grace\power noise data processing\11 Power noise&Voltage.xlsx')
    wb = openpyxl.load_workbook(file)

    sheet = wb['Noise&Voltage Test Data']
    # sheet.cell(row=6, column=10).value = data
    sheet.cell(row, col).value = data
    wb.save(r'E:\Grace\power noise data processing\11 Power noise&Voltage.xlsx')


def find_row(file, target):

    wb = openpyxl.load_workbook(file)
    sheet = wb['Noise&Voltage Test Data']


rootPath = os.getcwd()
# print(rootPath)

planPath = os.path.join(rootPath, 'test plan')
# print(planPath)

dataPath = os.path.join(rootPath, 'test data')
pnPath = os.path.join(dataPath, 'Power noise')
item_list = os.listdir(pnPath)

for item_id in item_list:

    print(item_id)
    idPath = os.path.join(pnPath, item_id)
    # print(os.listdir(idPath))

    dataName = os.listdir(idPath)
    # print(dataName)
    L_V_path = ''

    for data in dataName:
        if 'L_V.csv' in data:
            print(data)
            L_V_path = os.path.join(idPath, data)
            print(L_V_path)

        if 'H_V.csv' in data:
            print(data)
            H_V_path = os.path.join(idPath, data)
            print(H_V_path)

        if 'L_N.csv' in data:
            print(data)
            L_N_path = os.path.join(idPath, data)
            print(L_N_path)

        if 'H_N.csv' in data:
            print(data)
            H_N_path = os.path.join(idPath, data)
            print(H_N_path)

        csv_handle(L_V_path, 'RMS', 'Mean\'')
